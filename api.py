from flask import Blueprint, request, jsonify, current_app, send_file
from flask_login import login_required, current_user
from models import db, User, Sales, Returns, Target, Product, ActivityLog, UserRole, Department, DepartmentPermission, Task, TaskComment, Notification, Planning, PlanningSnapshot
from auth import (
    admin_required,
    representative_required,
    log_activity,
    permission_required,
    admin_or_department_manager_required,
    get_scoped_user_ids,
    is_user_in_scope,
)
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from sqlalchemy import func, and_, or_
import io
import base64
from werkzeug.utils import secure_filename
import os
from werkzeug.security import generate_password_hash

api = Blueprint('api', __name__)

# Timezone helpers (Turkey time)
TZ_TR = ZoneInfo('Europe/Istanbul')
def now_tr() -> datetime:
    return datetime.now(TZ_TR)

def today_tr() -> date:
    return now_tr().date()

# Helpers
def same_department(user_a: User, user_b: User) -> bool:
    try:
        return bool(user_a and user_b and user_a.department_id and user_b.department_id and user_a.department_id == user_b.department_id)
    except Exception:
        return False

def task_occurs_on(task: Task, day: date) -> bool:
    """Returns True if task should be shown on given day considering recurrence.
    - Non recurring: start_date == day or due_date == day
    - Recurring (anchor is start_date if provided, else due_date, else created_at.date):
        daily: every day in [anchor, due]
        weekly: same weekday as anchor in [anchor, due]
        monthly: same day-of-month as anchor in [anchor, due]
        yearly: same month/day as anchor in [anchor, due]
    If due_date is None for recurring, consider open-ended and only check lower bound.
    """
    if not task:
        return False
    # Non recurring
    if not task.is_recurring or (getattr(task, 'recurrence', 'none') or 'none') in ['none', '', None]:
        return (task.start_date == day) or (task.due_date == day)

    # Recurring
    anchor = task.start_date or task.due_date or (task.created_at.date() if task.created_at else None)
    if not anchor:
        return False
    if day < anchor:
        return False
    if task.due_date and day > task.due_date:
        return False
    pattern = (task.recurrence or '').lower()
    if pattern == 'daily':
        return True
    if pattern == 'weekly':
        return day.weekday() == anchor.weekday()
    if pattern == 'monthly':
        return day.day == anchor.day
    if pattern == 'yearly':
        return (day.month == anchor.month) and (day.day == anchor.day)
    return False

# Planlama API'leri (TR saati bazlı)
@api.route('/planning/today', methods=['GET', 'POST'])
@login_required
def planning_today():
    """Kullanıcının bugünkü plan kaydı: GET getirir, POST kaydeder (24 saat içinde düzenlenir).
    Not: Birden fazla öğeyi desteklemek için frontend çok satır girer; burada tek kayıt içinde saklanır.
    """
    user_id = current_user.id
    today = today_tr()
    plan = Planning.query.filter_by(representative_id=user_id, date=today).first()
    if request.method == 'GET':
        snapshots = PlanningSnapshot.query.filter_by(representative_id=user_id, date=today).order_by(PlanningSnapshot.created_at.asc()).all()
        return jsonify({
            'success': True,
            'plan': ({
                'id': plan.id,
                'date': plan.date.isoformat(),
                'yesterday_activities': plan.yesterday_activities,
                'today_plan': plan.today_plan,
                'challenges': plan.challenges,
                'created_at': plan.created_at.isoformat()
            } if plan else None),
            'snapshots': [{
                'id': s.id,
                'yesterday_activities': s.yesterday_activities,
                'today_plan': s.today_plan,
                'challenges': s.challenges,
                'created_at': s.created_at.isoformat(),
            } for s in snapshots]
        })
    data = request.get_json() or {}
    if not plan:
        plan = Planning(
            representative_id=user_id,
            date=today,
            yesterday_activities=data.get('yesterday_activities'),
            today_plan=data.get('today_plan'),
            challenges=data.get('challenges')
        )
        db.session.add(plan)
    else:
        if not plan.can_edit():
            return jsonify({'success': False, 'error': '24 saat geçti, düzenleme yapılamaz'}), 400
        plan.yesterday_activities = data.get('yesterday_activities')
        plan.today_plan = data.get('today_plan')
        plan.challenges = data.get('challenges')
    # snapshot kaydet
    snap = PlanningSnapshot(
        representative_id=user_id,
        date=today,
        yesterday_activities=plan.yesterday_activities,
        today_plan=plan.today_plan,
        challenges=plan.challenges,
    )
    db.session.add(snap)
    db.session.commit()
    return jsonify({'success': True})

@api.route('/tasks/<int:task_id>', methods=['DELETE'])
@login_required
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)
    # Only admin can delete any task; DM can delete tasks in their department they manage; users cannot delete
    if current_user.is_admin():
        pass
    elif current_user.is_department_manager():
        # DM: Kendi departmanına ait tüm görevleri silebilir.
        # Kapsam şunları içerir:
        # - task.department_id DM'nin departmanı
        # - assigned_to DM'nin departmanında
        # - created_by DM'nin departmanında
        try:
            assignee_dept = getattr(getattr(task, 'assigned_to', None), 'department_id', None)
        except Exception:
            assignee_dept = None
        try:
            creator_dept = getattr(getattr(task, 'created_by', None), 'department_id', None)
        except Exception:
            creator_dept = None
        if not (
            task.department_id == current_user.department_id or
            assignee_dept == current_user.department_id or
            creator_dept == current_user.department_id
        ):
            return jsonify({'success': False, 'error': 'Silme yetkiniz yok'}), 403
    else:
        return jsonify({'success': False, 'error': 'Silme yetkiniz yok'}), 403
    try:
        db.session.delete(task)
        db.session.commit()
        # Best-effort: cache/summary tutarlı olsun diye bir şey döndür
        return jsonify({'success': True, 'deleted_id': task_id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/planning/months', methods=['GET'])
@login_required
def planning_months():
    """Belirli bir yıl için 12 ayı klasör mantığında döner.
    Admin/DM: user_id parametresi ile başka kullanıcıyı görüntüleyebilir.
    Parametreler: year (int) – verilmezse TR güncel yıl kullanılır.
    """
    requested_user_id = request.args.get('user_id', type=int)
    # Hedef kullanıcı
    if current_user.is_admin():
        user_id = requested_user_id or current_user.id
    elif current_user.is_department_manager():
        user_id = requested_user_id if (requested_user_id and is_user_in_scope(requested_user_id)) else current_user.id
    else:
        user_id = current_user.id
    # Hedef yıl
    param_year = request.args.get('year', type=int)
    base_date = today_tr()
    year = param_year or base_date.year
    months = []
    for m in range(1, 13):
        y = year
        start = date(y, m, 1)
        end_month = m + 1
        end_year = y + (1 if end_month == 13 else 0)
        end = date(end_year, 1 if end_month == 13 else end_month, 1)
        days_with_entries = Planning.query.filter(
            Planning.representative_id == user_id,
            Planning.date >= start,
            Planning.date < end
        ).count()
        months.append({
            'year': y,
            'month': m,
            'label': f"{y}-{str(m).zfill(2)}",
            'days_with_entries': days_with_entries
        })
    return jsonify({'success': True, 'year': year, 'months': months})

@api.route('/planning/month', methods=['GET'])
@login_required
def planning_month_detail():
    """Ay takvimi: günlerde plan var mı ve görev var mı bilgisi.
    Admin/DM: user_id parametresi ile başka kullanıcıyı görüntüleyebilir.
    """
    requested_user_id = request.args.get('user_id', type=int)
    if current_user.is_admin():
        user_id = requested_user_id or current_user.id
    elif current_user.is_department_manager():
        user_id = requested_user_id if (requested_user_id and is_user_in_scope(requested_user_id)) else current_user.id
    else:
        user_id = current_user.id
    param_year = request.args.get('year', type=int)
    param_month = request.args.get('month', type=int)
    today = today_tr()
    y = param_year or today.year
    m = param_month or today.month
    start = date(y, m, 1)
    end_month = m + 1
    end_year = y + (1 if end_month == 13 else 0)
    end = date(end_year, 1 if end_month == 13 else end_month, 1)

    # Plan var mı (ana plan veya snapshot)
    plan_days = set()
    plan_days.update(d[0] for d in db.session.query(Planning.date).filter(
        Planning.representative_id == user_id,
        Planning.date >= start,
        Planning.date < end
    ).all())
    # Snapshot'lar da plan göstergesine dahil
    plan_days.update(d[0] for d in db.session.query(PlanningSnapshot.date).filter(
        PlanningSnapshot.representative_id == user_id,
        PlanningSnapshot.date >= start,
        PlanningSnapshot.date < end
    ).all())
    # Görev var mı: o ayın herhangi bir gününde occurrence var mı (recurrence dahil)
    task_days = set()
    task_start_days = set()
    task_due_days = set()
    
    tasks = Task.query.filter(
        (Task.assigned_to_id == user_id) | (Task.created_by_id == user_id)
    ).all()
    from calendar import monthrange
    num_days = monthrange(y, m)[1]

    # Ayrıntı haritası: sadece admin veya departman yöneticisi için küçük özet döneceğiz
    is_privileged = current_user.is_admin() or current_user.is_department_manager()
    day_tasks_map = {}
    for dnum in range(1, num_days+1):
        dte = date(y, m, dnum)
        for t in tasks:
            if task_occurs_on(t, dte):
                task_days.add(dte)
                if is_privileged:
                    lst = day_tasks_map.setdefault(dte, [])
                    # İsimleri hazırla
                    try:
                        to_name = (t.assigned_to.get_full_name() if getattr(t, 'assigned_to', None) else None)
                    except Exception:
                        to_name = None
                    try:
                        by_name = (t.assigned_by.get_full_name() if getattr(t, 'assigned_by', None) else None)
                    except Exception:
                        by_name = None
                    # Geriye dönük: assigned_by yoksa created_by kullan
                    if not by_name:
                        try:
                            by_name = (t.created_by.get_full_name() if getattr(t, 'created_by', None) else None)
                        except Exception:
                            by_name = None
                    lst.append({
                        'assigned_to_name': to_name or '-',
                        'assigned_by_name': by_name or '-',
                    })

    # Atama ve bitiş işaretleri
    for t in tasks:
        try:
            # Atama günü: start_date varsa onu, yoksa created_at tarihini kullan
            if t.start_date and (t.start_date >= start and t.start_date < end):
                task_start_days.add(t.start_date)
            elif getattr(t, 'created_at', None):
                ca = t.created_at.date() if isinstance(t.created_at, datetime) else None
                if ca and (ca >= start and ca < end):
                    task_start_days.add(ca)
        except Exception:
            pass
        try:
            if t.due_date and (t.due_date >= start and t.due_date < end):
                task_due_days.add(t.due_date)
        except Exception:
            pass
    days = []
    from calendar import monthrange
    num_days = monthrange(y, m)[1]
    for dnum in range(1, num_days+1):
        dte = date(y, m, dnum)
        item = {
            'day': dnum,
            'has_planning': dte in plan_days,
            'has_tasks': dte in task_days,
            'has_task_assign': dte in task_start_days,
            'has_task_due': dte in task_due_days
        }
        if is_privileged and dte in day_tasks_map:
            # İlk 2 görevin atama özetini gönder, kalan sayısını belirt
            meta = day_tasks_map.get(dte, [])
            item['tasks_meta'] = meta[:2]
            if len(meta) > 2:
                item['tasks_meta_more'] = len(meta) - 2
        days.append(item)
    return jsonify({'success': True, 'year': y, 'month': m, 'days': days})

@api.route('/planning/years', methods=['GET'])
@login_required
def planning_years():
    """Kullanıcı için yıl klasörleri. Admin/DM user_id ile başkasını görüntüleyebilir."""
    requested_user_id = request.args.get('user_id', type=int)
    if current_user.is_admin():
        user_id = requested_user_id or current_user.id
    elif current_user.is_department_manager():
        user_id = requested_user_id if (requested_user_id and is_user_in_scope(requested_user_id)) else current_user.id
    else:
        user_id = current_user.id

    # Planning yılları
    planning_years_rows = db.session.query(func.strftime('%Y', Planning.date)).filter(
        Planning.representative_id == user_id
    ).distinct().all()
    years_set = set(int(y[0]) for y in planning_years_rows if y and y[0])

    # Task yılları (due_date)
    task_years_rows = db.session.query(func.strftime('%Y', Task.due_date)).filter(
        ((Task.assigned_to_id == user_id) | (Task.created_by_id == user_id)) & (Task.due_date.isnot(None))
    ).distinct().all()
    years_set.update(int(y[0]) for y in task_years_rows if y and y[0])

    # En azından içinde bulunduğumuz yılı ekle
    years_set.add(today_tr().year)
    years = sorted(list(years_set), reverse=True)

    result = []
    for y in years:
        start = date(y, 1, 1)
        end = date(y + 1, 1, 1)
        days_with_entries = Planning.query.filter(
            Planning.representative_id == user_id,
            Planning.date >= start,
            Planning.date < end
        ).count()
        result.append({'year': y, 'label': f"{y}", 'days_with_entries': days_with_entries})

    return jsonify({'success': True, 'years': result})

@api.route('/planning/day', methods=['GET', 'DELETE'])
@login_required
def planning_day_detail():
    """Belirli gün için plan ve görev detayları.
    Admin/DM: user_id parametresi ile başka kullanıcıyı görüntüleyebilir.
    """
    requested_user_id = request.args.get('user_id', type=int)
    if current_user.is_admin():
        user_id = requested_user_id or current_user.id
    elif current_user.is_department_manager():
        user_id = requested_user_id if (requested_user_id and is_user_in_scope(requested_user_id)) else current_user.id
    else:
        user_id = current_user.id
    date_str = request.args.get('date')
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'error': 'Geçersiz tarih formatı'}), 400
    else:
        target_date = today_tr()

    if request.method == 'DELETE':
        # Yalnız admin silebilir (başka kullanıcının kaydı için user_id zorunlu)
        if not current_user.is_admin():
            return jsonify({'success': False, 'error': 'Silme yetkiniz yok'}), 403
        if not requested_user_id:
            # Admin başka kullanıcının planını siliyorsa user_id gönderilmeli
            requested_user_id = current_user.id
            if user_id != current_user.id:
                return jsonify({'success': False, 'error': 'admin olarak başka bir kullanıcı için user_id parametresi gerekli'}), 400
        # Delete main plan and all snapshots for that date/user
        deleted = 0
        plan_q = Planning.query.filter_by(representative_id=user_id, date=target_date)
        plan = plan_q.first()
        if plan:
            db.session.delete(plan)
            deleted += 1
        snaps = PlanningSnapshot.query.filter_by(representative_id=user_id, date=target_date).all()
        for s in snaps:
            db.session.delete(s)
        deleted += len(snaps)
        db.session.commit()
        return jsonify({'success': True, 'deleted': deleted})

    plan = Planning.query.filter_by(representative_id=user_id, date=target_date).first()
    snapshots = PlanningSnapshot.query.filter_by(representative_id=user_id, date=target_date).order_by(PlanningSnapshot.created_at.asc()).all()
    tasks = Task.query.filter(
        (Task.assigned_to_id == user_id) | (Task.created_by_id == user_id)
    ).all()
    # Recurrence-aware filter
    tasks = [t for t in tasks if task_occurs_on(t, target_date)]
    # Sort by priority (high > normal > low), then due_date
    priority_order = {'high': 0, 'normal': 1, 'low': 2}
    tasks.sort(key=lambda t: (priority_order.get((t.priority or 'normal').lower(), 1), (t.due_date or date.max)))

    return jsonify({
        'success': True,
        'date': target_date.isoformat(),
        'plan': ({
            'id': plan.id,
            'yesterday_activities': plan.yesterday_activities,
            'today_plan': plan.today_plan,
            'challenges': plan.challenges,
            'created_at': plan.created_at.isoformat()
        } if plan else None),
        'snapshots': [{
            'id': s.id,
            'yesterday_activities': s.yesterday_activities,
            'today_plan': s.today_plan,
            'challenges': s.challenges,
            'created_at': s.created_at.isoformat(),
        } for s in snapshots],
        'tasks': [{
            'id': t.id,
            'title': t.title,
            'status': t.status,
            'priority': t.priority,
            'due_date': t.due_date.isoformat() if t.due_date else None,
            'created_at': t.created_at.isoformat() if t.created_at else None,
            'start_date': t.start_date.isoformat() if t.start_date else None,
            'description': t.description
        } for t in tasks]
    })

@api.route('/planning/archive/departments', methods=['GET'])
@login_required
def planning_archive_departments():
    """Departman -> kullanıcılar ve kullanıcı bazlı görev istatistikleri."""
    scoped_ids = get_scoped_user_ids()
    # Opsiyonel departman filtresi
    requested_dept_id = request.args.get('department_id', type=int)
    # Departmanlar
    departments_q = Department.query
    if scoped_ids is not None and not current_user.is_admin():
        # DM: kendi departmanı
        departments_q = departments_q.filter(Department.id == current_user.department_id)
    if requested_dept_id:
        departments_q = departments_q.filter(Department.id == requested_dept_id)
    departments = departments_q.all()
    today = today_tr()

    result = []
    for dept in departments:
        users_q = User.query.filter(User.department_id == dept.id)
        if scoped_ids is not None:
            users_q = users_q.filter(User.id.in_(scoped_ids))
        dept_users = users_q.all()
        users_payload = []
        for u in dept_users:
            total_tasks = Task.query.filter_by(assigned_to_id=u.id).count()
            completed_tasks = Task.query.filter_by(assigned_to_id=u.id, status='completed').count()
            in_progress = Task.query.filter_by(assigned_to_id=u.id, status='in_progress').count()
            pending = Task.query.filter(Task.assigned_to_id==u.id, Task.status.in_(['pending','requested'])).count()
            overdue = Task.query.filter(
                Task.assigned_to_id==u.id,
                Task.due_date.isnot(None),
                Task.due_date < today,
                Task.status != 'completed',
                Task.status != 'cancelled'
            ).count()
            completion_rate = (completed_tasks/total_tasks*100) if total_tasks>0 else 0
            users_payload.append({
                'id': u.id,
                'name': u.get_full_name(),
                'username': u.username,
                'stats': {
                    'total': total_tasks,
                    'completed': completed_tasks,
                    'in_progress': in_progress,
                    'pending': pending,
                    'overdue': overdue,
                    'completion_rate': completion_rate
                }
            })
        result.append({
            'id': dept.id,
            'name': dept.name,
            'users': users_payload
        })
    return jsonify({'success': True, 'departments': result})

@api.route('/departments/simple', methods=['GET'])
@login_required
def list_simple_departments():
    """Basit departman listesi (id, name). Admin: tümü; DM/Kullanıcı: kendi departmanı."""
    if current_user.is_admin():
        depts = Department.query.order_by(Department.name.asc()).all()
    else:
        if current_user.department_id:
            depts = Department.query.filter(Department.id == current_user.department_id).all()
        else:
            depts = []
    return jsonify({'success': True, 'departments': [{'id': d.id, 'name': d.name} for d in depts]})

# Görev Yönetimi API
@api.route('/tasks', methods=['GET'])
@login_required
def list_tasks():
    """Görevleri erişim kapsamına göre listele.
    - Admin: tüm görevler
    - DM: kendi departmanı (department_id eşleşen) ve kendisinin oluşturduğu/görevlendirildiği görevler
    - Kullanıcı: kendisine atanan, kendisinin yarattığı ve departmanına ait (görme izni varsa) görevler
    Filtreler: status, assigned_to_id, start_date, end_date (due_date için)
    """
    # Çoklu durum filtresi destekle: status birden fazla olabilir veya virgül-separe
    raw_status_list = request.args.getlist('status') or []
    if not raw_status_list:
        single_status = request.args.get('status')
        if single_status:
            raw_status_list = [s for s in single_status.split(',') if s]
    statuses = [s.strip() for s in raw_status_list if s and isinstance(s, str)]
    assigned_to_id = request.args.get('assigned_to_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    query = Task.query

    if current_user.is_admin():
        pass
    elif current_user.is_department_manager():
        # DM: departman görevleri + kendi oluşturdukları/atandıkları
        query = query.filter(
            (Task.department_id == current_user.department_id) |
            (Task.assigned_to_id == current_user.id) |
            (Task.created_by_id == current_user.id)
        )
    else:
        # Kullanıcı: kendisiyle ilgili görevler
        query = query.filter(
            (Task.assigned_to_id == current_user.id) |
            (Task.created_by_id == current_user.id)
        )

    if statuses:
        query = query.filter(Task.status.in_(statuses))
    if assigned_to_id:
        try:
            assigned_to_id = int(assigned_to_id)
        except Exception:
            return jsonify({'success': False, 'error': 'assigned_to_id geçersiz'}), 400
        query = query.filter(Task.assigned_to_id == assigned_to_id)
    # Not: listedeki filtreler due_date üzerineydi; ek olarak start_date'i de kapsayacağız
    if start_date:
        sd = datetime.strptime(start_date, '%Y-%m-%d').date()
        query = query.filter(or_(Task.due_date == None, Task.due_date >= sd))
    if end_date:
        ed = datetime.strptime(end_date, '%Y-%m-%d').date()
        query = query.filter(or_(Task.due_date == None, Task.due_date <= ed))

    tasks = query.order_by(Task.due_date.is_(None), Task.due_date.asc(), Task.created_at.desc()).all()
    return jsonify({
        'success': True,
        'tasks': [
            {
                'id': t.id,
                'title': t.title,
                'description': t.description,
                'department_id': t.department_id,
                'assigned_by_id': t.assigned_by_id,
                'assigned_to_id': t.assigned_to_id,
                'created_by_id': t.created_by_id,
                'status': t.status,
                'priority': t.priority,
                'due_date': t.due_date.isoformat() if t.due_date else None,
                'start_date': t.start_date.isoformat() if t.start_date else None,
                'is_recurring': t.is_recurring,
                'recurrence': t.recurrence,
                'created_at': t.created_at.isoformat(),
                'updated_at': t.updated_at.isoformat(),
            } for t in tasks
        ]
    })

@api.route('/tasks', methods=['POST'])
@login_required
def create_task():
    """Görev oluştur.
    - Admin: herkese görev verebilir
    - DM: sadece kendi departmanındaki kullanıcılara görev verebilir
    - Kullanıcı: "requested" statüsünde yöneticisine iş isteği açabilir (assigned_to_id boş bırakılabilir veya yöneticisine atanır)
    """
    data = request.get_json() or {}
    title = data.get('title')
    if not title:
        return jsonify({'success': False, 'error': 'Başlık zorunludur'}), 400

    # Çoklu atama desteği: assigned_to_ids varsa döngü ile birden fazla görev
    assigned_to_id = data.get('assigned_to_id')
    assigned_to_ids = data.get('assigned_to_ids') or []
    if isinstance(assigned_to_ids, list):
        try:
            assigned_to_ids = [int(x) for x in assigned_to_ids if x is not None and x != ""]
        except Exception:
            return jsonify({'success': False, 'error': 'assigned_to_ids geçersiz'}), 400
    try:
        if assigned_to_id is not None and assigned_to_id != "":
            assigned_to_id = int(assigned_to_id)
        else:
            assigned_to_id = None
    except Exception:
        return jsonify({'success': False, 'error': 'assigned_to_id geçersiz'}), 400
    due_date = data.get('due_date')
    start_date = data.get('start_date')
    description = data.get('description')
    priority = data.get('priority', 'normal')
    is_recurring = data.get('is_recurring', False)
    recurrence = data.get('recurrence', 'none')

    # Yardımcı: tek görev oluşturma
    def build_and_persist_task(assignee_id):
        task = Task(
            title=title,
            description=description,
            department_id=current_user.department_id,
            created_by_id=current_user.id,
            priority=priority,
            is_recurring=is_recurring,
            recurrence=recurrence,
        )
        if start_date:
            task.start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if due_date:
            task.due_date = datetime.strptime(due_date, '%Y-%m-%d').date()

        # Policy update: everyone can assign tasks to anyone
        # Normalize assignee
        if assignee_id:
            assignee = User.query.get(assignee_id)
            if not assignee:
                raise ValueError('Atanacak kullanıcı bulunamadı')
        # Status and assignment
        task.status = 'pending'
        task.assigned_by_id = current_user.id
        task.assigned_to_id = assignee_id
        db.session.add(task)
        db.session.flush()
        return task

    created_tasks = []
    try:
        targets = assigned_to_ids if assigned_to_ids else [assigned_to_id]
        if not targets:
            targets = [None]
        for tid in targets:
            created_tasks.append(build_and_persist_task(tid))
        db.session.commit()
    except PermissionError as pe:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(pe)}), 403
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

    # Bildirimler: atanan kullanıcı(lar)a, DM'ye ve adminlere haber ver
    try:
        # Assigned user notification
        for task in created_tasks:
            if task.assigned_to_id:
                db.session.add(Notification(
                    to_user_id=task.assigned_to_id,
                    created_by_id=current_user.id,
                    title='Yeni Görev',
                    message=f"Size yeni bir görev atandı: {title}",
                    url='/tasks',
                    entity_type='task',
                    entity_id=task.id
                ))
        # Department manager notification
        if current_user.department_id:
            manager = User.query.filter_by(role=UserRole.DEPARTMENT_MANAGER, department_id=current_user.department_id).first()
            if manager and manager.id != current_user.id:
                db.session.add(Notification(
                    to_user_id=manager.id,
                    created_by_id=current_user.id,
                    title='Yeni Görev Oluşturuldu',
                    message=f"{current_user.get_full_name()} yeni görev oluşturdu: {title}",
                    url='/tasks',
                    entity_type='task',
                    entity_id=task.id
                ))
        # Admin notifications
        admins = User.query.filter_by(role=UserRole.ADMIN).all()
        for adm in admins:
            if adm.id != current_user.id:
                db.session.add(Notification(
                    to_user_id=adm.id,
                    created_by_id=current_user.id,
                    title='Görev Etkinliği',
                    message=f"{current_user.get_full_name()} görev oluşturdu: {title}",
                    url='/tasks',
                    entity_type='task',
                    entity_id=task.id
                ))
        db.session.commit()
    except Exception:
        db.session.rollback()

    # Çoklu görev durumunda tüm id'leri döndür
    return jsonify({'success': True, 'task_ids': [t.id for t in created_tasks]}), 201

@api.route('/tasks/<int:task_id>', methods=['GET', 'PUT'])
@login_required
def get_or_update_task(task_id):
    task = Task.query.get_or_404(task_id)
    # Erişim kontrolü
    if not (current_user.is_admin() or is_user_in_scope(task.assigned_to_id or current_user.id) or is_user_in_scope(task.created_by_id)):
        return jsonify({'success': False, 'error': 'Göreve erişim yetkiniz yok'}), 403

    if request.method == 'GET':
        return jsonify({
            'success': True,
            'task': {
                'id': task.id,
                'title': task.title,
                'description': task.description,
                'department_id': task.department_id,
                'assigned_by_id': task.assigned_by_id,
                'assigned_to_id': task.assigned_to_id,
                'created_by_id': task.created_by_id,
                'status': task.status,
                'priority': task.priority,
                'due_date': task.due_date.isoformat() if task.due_date else None,
                'start_date': task.start_date.isoformat() if task.start_date else None,
                'is_recurring': task.is_recurring,
                'recurrence': task.recurrence,
                'created_at': task.created_at.isoformat(),
                'updated_at': task.updated_at.isoformat(),
            }
        })

    # PUT: durum/atanan/gün sınırı vb güncelleme
    data = request.get_json() or {}

    # Admin veya DM görev üzerinde tam yetkili; kullanıcı sadece kendi görevinde durum/progress güncelleyebilir
    if current_user.is_admin() or current_user.is_department_manager():
        if 'status' in data:
            task.status = data['status']
        if 'assigned_to_id' in data:
            new_assignee = data['assigned_to_id']
            try:
                if new_assignee is not None and new_assignee != "":
                    new_assignee = int(new_assignee)
                else:
                    new_assignee = None
            except Exception:
                return jsonify({'success': False, 'error': 'assigned_to_id geçersiz'}), 400
            if new_assignee and not is_user_in_scope(new_assignee):
                return jsonify({'success': False, 'error': 'Bu kullanıcıya atama yetkiniz yok'}), 403
            task.assigned_to_id = new_assignee
        if 'due_date' in data:
            task.due_date = datetime.strptime(data['due_date'], '%Y-%m-%d').date() if data['due_date'] else None
        if 'start_date' in data:
            task.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data['start_date'] else None
        if 'title' in data:
            task.title = data['title']
        if 'description' in data:
            task.description = data['description']
        if 'priority' in data:
            task.priority = data['priority']
        if 'is_recurring' in data:
            task.is_recurring = bool(data['is_recurring'])
        if 'recurrence' in data:
            task.recurrence = data['recurrence'] or 'none'
    else:
        # Kullanıcı: sadece kendi görevinde status/progress
        if task.assigned_to_id != current_user.id and task.created_by_id != current_user.id:
            return jsonify({'success': False, 'error': 'Görevi güncelleme yetkiniz yok'}), 403
        if 'status' in data:
            # Kullanıcı "requested" -> "pending" yapamaz; admin/DM onaylar
            allowed_status = {'in_progress', 'completed', 'cancelled'}
            if data['status'] not in allowed_status:
                return jsonify({'success': False, 'error': 'Geçersiz durum güncellemesi'}), 400
            task.status = data['status']

    db.session.commit()
    return jsonify({'success': True}), 200

@api.route('/tasks/<int:task_id>/approve', methods=['POST'])
@login_required
def approve_task(task_id):
    """Atanan kullanıcının görevi onaylaması (pending/requested -> in_progress).
    Bildirim: oluşturana, departman yöneticisine ve adminlere gider.
    """
    task = Task.query.get_or_404(task_id)
    # Sadece atanan kullanıcı onaylayabilir
    if not task.assigned_to_id or task.assigned_to_id != current_user.id:
        return jsonify({'success': False, 'error': 'Bu görevi onaylama yetkiniz yok'}), 403
    # Uygun durumlar
    if (task.status or '').lower() not in ['pending', 'requested']:
        return jsonify({'success': False, 'error': 'Bu görev bu durumda onaylanamaz'}), 400

    try:
        task.status = 'in_progress'
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

    # Onay bildirimi gönder
    try:
        notify_user_ids = set()
        if task.created_by_id:
            notify_user_ids.add(task.created_by_id)
        if task.department_id:
            dm = User.query.filter_by(role=UserRole.DEPARTMENT_MANAGER, department_id=task.department_id).first()
            if dm:
                notify_user_ids.add(dm.id)
        # Adminlere de bilgi ver
        admins = User.query.filter_by(role=UserRole.ADMIN).all()
        for adm in admins:
            notify_user_ids.add(adm.id)
        # Kendine bildirim gönderme
        notify_user_ids.discard(current_user.id)

        for uid in notify_user_ids:
            db.session.add(Notification(
                to_user_id=uid,
                created_by_id=current_user.id,
                title='Görev Onayı',
                message=f"{current_user.get_full_name()} bir görevi onayladı: {task.title}",
                url='/tasks',
                entity_type='task',
                entity_id=task.id
            ))
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'success': True}), 200

@api.route('/tasks/<int:task_id>/deliver', methods=['POST'])
@login_required
def deliver_task(task_id):
    """Atanan kullanıcının görevi teslim etmesi (tamamlandı). Bildirim gönderilir."""
    task = Task.query.get_or_404(task_id)
    # Sadece atanan kullanıcı teslim edebilir
    if not task.assigned_to_id or task.assigned_to_id != current_user.id:
        return jsonify({'success': False, 'error': 'Bu görevi teslim etme yetkiniz yok'}), 403
    # Uygun durumlar: in_progress, pending, requested
    if (task.status or '').lower() not in ['in_progress', 'pending', 'requested']:
        return jsonify({'success': False, 'error': 'Bu görev bu durumda teslim edilemez'}), 400
    # Son tarih geçtiyse ve tamamlanmadıysa teslim edilemez
    try:
        if task.due_date is not None:
            from datetime import datetime
            now = datetime.utcnow()
            # due_date'in timezone bilgisi yoksa doğrudan karşılaştır
            if task.due_date < now:
                return jsonify({'success': False, 'error': 'Görev geciktiği için teslim edilemez'}), 400
    except Exception:
        pass

    try:
        task.status = 'completed'
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

    # Teslim bildirimi gönder
    try:
        notify_user_ids = set()
        if task.created_by_id:
            notify_user_ids.add(task.created_by_id)
        if task.department_id:
            dm = User.query.filter_by(role=UserRole.DEPARTMENT_MANAGER, department_id=task.department_id).first()
            if dm:
                notify_user_ids.add(dm.id)
        admins = User.query.filter_by(role=UserRole.ADMIN).all()
        for adm in admins:
            notify_user_ids.add(adm.id)
        notify_user_ids.discard(current_user.id)
        for uid in notify_user_ids:
            db.session.add(Notification(
                to_user_id=uid,
                created_by_id=current_user.id,
                title='Görev Teslim Edildi',
                message=f"{current_user.get_full_name()} görevi teslim etti: {task.title}",
                url='/tasks',
                entity_type='task',
                entity_id=task.id
            ))
        db.session.commit()
    except Exception:
        db.session.rollback()

    return jsonify({'success': True}), 200



@api.route('/tasks/<int:task_id>/comments', methods=['GET', 'POST'])
@login_required
def task_comments(task_id):
    task = Task.query.get_or_404(task_id)
    if not (current_user.is_admin() or is_user_in_scope(task.assigned_to_id or current_user.id) or is_user_in_scope(task.created_by_id)):
        return jsonify({'success': False, 'error': 'Göreve erişim yetkiniz yok'}), 403

    if request.method == 'GET':
        comments = TaskComment.query.filter_by(task_id=task_id).order_by(TaskComment.created_at.asc()).all()
        return jsonify({
            'success': True,
            'comments': [{
                'id': c.id,
                'task_id': c.task_id,
                'user_id': c.user_id,
                'user_name': c.user.get_full_name() if c.user else '-',
                'comment': c.comment,
                'created_at': c.created_at.isoformat(),
            } for c in comments]
        })

    data = request.get_json() or {}
    text = data.get('comment')
    if not text:
        return jsonify({'success': False, 'error': 'Yorum zorunludur'}), 400
    c = TaskComment(task_id=task_id, user_id=current_user.id, comment=text)
    db.session.add(c)
    db.session.commit()
    # Yorum bildirimi gönder
    try:
        notify_user_ids = set()
        if task.assigned_to_id:
            notify_user_ids.add(task.assigned_to_id)
        if task.created_by_id:
            notify_user_ids.add(task.created_by_id)
        if task.department_id:
            dm = User.query.filter_by(role=UserRole.DEPARTMENT_MANAGER, department_id=task.department_id).first()
            if dm:
                notify_user_ids.add(dm.id)
        # Adminler
        for adm in User.query.filter_by(role=UserRole.ADMIN).all():
            notify_user_ids.add(adm.id)
        # Kendine bildirim gönderme
        notify_user_ids.discard(current_user.id)
        for uid in notify_user_ids:
            db.session.add(Notification(
                to_user_id=uid,
                created_by_id=current_user.id,
                title='Görev Yorumu',
                message=f"{current_user.get_full_name()} bir yorum ekledi.",
                url='/tasks',
                entity_type='task',
                entity_id=task_id
            ))
        db.session.commit()
    except Exception:
        db.session.rollback()
    return jsonify({'success': True, 'comment_id': c.id}), 201

# Bildirim API'leri
@api.route('/notifications', methods=['GET'])
@login_required
def list_notifications():
    """Giriş yapan kullanıcının bildirimlerini getir."""
    only_unread = request.args.get('unread', 'false').lower() == 'true'
    q = Notification.query.filter_by(to_user_id=current_user.id)
    if only_unread:
        q = q.filter_by(is_read=False)
    notifs = q.order_by(Notification.created_at.desc()).limit(50).all()
    return jsonify({
        'success': True,
        'notifications': [{
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'url': n.url,
            'entity_type': n.entity_type,
            'entity_id': n.entity_id,
            'is_read': n.is_read,
            'created_at': n.created_at.isoformat()
        } for n in notifs]
    })

@api.route('/tasks/due-soon', methods=['GET'])
@login_required
def tasks_due_soon():
    """Yaklaşan bitiş tarihli görevler (bugün ve sonraki 3 gün)."""
    try:
        days = request.args.get('days', 3, type=int)
        today = datetime.utcnow().date()
        until = today + timedelta(days=max(0, days))

        query = Task.query.filter(
            Task.due_date != None,
            Task.due_date >= today,
            Task.due_date <= until,
            Task.status.in_(['pending','in_progress','requested'])
        )

        # Erişim kapsamı: Admin tümü, DM departman + kendisi, kullanıcı kendiyle ilgili
        if current_user.is_admin():
            pass
        elif current_user.is_department_manager():
            query = query.filter(
                (Task.department_id == current_user.department_id) |
                (Task.assigned_to_id == current_user.id) |
                (Task.created_by_id == current_user.id)
            )
        else:
            query = query.filter(
                (Task.assigned_to_id == current_user.id) |
                (Task.created_by_id == current_user.id)
            )

        tasks = query.order_by(Task.due_date.asc()).limit(50).all()
        return jsonify({'success': True, 'tasks': [{
            'id': t.id,
            'title': t.title,
            'due_date': t.due_date.isoformat() if t.due_date else None,
            'start_date': t.start_date.isoformat() if t.start_date else None,
            'status': t.status,
            'priority': t.priority
        } for t in tasks]})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/notifications/unread-count', methods=['GET'])
@login_required
def unread_notifications_count():
    cnt = Notification.query.filter_by(to_user_id=current_user.id, is_read=False).count()
    return jsonify({'success': True, 'count': cnt})

@api.route('/notifications/<int:notif_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    notif = Notification.query.filter_by(id=notif_id, to_user_id=current_user.id).first_or_404()
    notif.is_read = True
    notif.read_at = datetime.utcnow()
    db.session.commit()

    # Okundu bildirimi: kullanıcının yöneticisine ve adminlere haber ver
    try:
        notify_user_ids = set()
        # Departman yöneticisi
        if current_user.department_id:
            dm = User.query.filter_by(role=UserRole.DEPARTMENT_MANAGER, department_id=current_user.department_id).first()
            if dm:
                notify_user_ids.add(dm.id)
        # Adminler
        for adm in User.query.filter_by(role=UserRole.ADMIN).all():
            notify_user_ids.add(adm.id)
        # Kendine bildirim gönderme
        notify_user_ids.discard(current_user.id)
        message = f"{current_user.get_full_name()} bir bildirimi görüntüledi."
        if notif.entity_type == 'task' and notif.entity_id:
            message = f"{current_user.get_full_name()} görev bildirimi görüntüledi (ID: {notif.entity_id})."
        for uid in notify_user_ids:
            db.session.add(Notification(
                to_user_id=uid,
                created_by_id=current_user.id,
                title='Bildirim Görüntülendi',
                message=message,
                url=notif.url or '/tasks',
                entity_type=notif.entity_type,
                entity_id=notif.entity_id
            ))
        db.session.commit()
    except Exception:
        db.session.rollback()
    return jsonify({'success': True})

# Hedef yönetimi
@api.route('/targets', methods=['GET'])
@login_required
def get_targets():
    """Kullanıcının hedeflerini getir"""
    if current_user.is_admin():
        # Admin tüm hedefleri görebilir
        targets = Target.query.all()
    else:
        # Diğer kullanıcılar sadece kendi hedeflerini görebilir
        targets = Target.query.filter_by(user_id=current_user.id).all()
    
    return jsonify({
        'targets': [{
            'id': target.id,
            'user_id': target.user_id,
            'year': target.year,
            'month': target.month,
            'target_amount': target.target_amount,
            'created_at': target.created_at.isoformat(),
            'updated_at': target.updated_at.isoformat()
        } for target in targets]
    }), 200

@api.route('/targets', methods=['POST'])
@admin_or_department_manager_required
def create_target():
    """Admin veya Departman Yöneticisi: Kullanıcıya hedef ata (DM sadece kendi departmanındaki kullanıcılar için)."""
    data = request.get_json()
    
    required_fields = ['user_id', 'year', 'month', 'target_amount']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'{field} alanı gerekli'}), 400
    
    # Kullanıcı erişim kontrolü (DM kendi departmanındaki kullanıcıya, Admin herkese)
    user = User.query.get(data['user_id'])
    if not user:
        return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
    if not is_user_in_scope(user.id):
        return jsonify({'error': 'Bu kullanıcıya hedef atamak için yetkiniz yok'}), 403
    
    # Aynı ay için hedef var mı kontrol et
    existing_target = Target.query.filter_by(
        user_id=data['user_id'],
        year=data['year'],
        month=data['month']
    ).first()
    
    if existing_target:
        return jsonify({'error': 'Bu ay için zaten hedef tanımlanmış'}), 400
    
    target = Target(
        user_id=data['user_id'],
        year=data['year'],
        month=data['month'],
        target_amount=data['target_amount']
    )
    
    db.session.add(target)
    db.session.commit()
    
    log_activity('target_create', f'Hedef oluşturuldu: {user.username} - {data["year"]}/{data["month"]}')
    
    return jsonify({'message': 'Hedef oluşturuldu'}), 201

@api.route('/targets/<int:target_id>', methods=['PUT'])
@admin_or_department_manager_required
def update_target(target_id):
    """Admin: Hedef güncelle"""
    target = Target.query.get_or_404(target_id)
    data = request.get_json()
    
    if 'target_amount' in data:
        target.target_amount = data['target_amount']
    
    db.session.commit()
    
    log_activity('target_update', f'Hedef güncellendi: ID {target_id}')
    
    return jsonify({'message': 'Hedef güncellendi'}), 200

@api.route('/targets/<int:target_id>', methods=['DELETE'])
@admin_or_department_manager_required
def delete_target(target_id):
    """Admin: Hedef sil"""
    target = Target.query.get_or_404(target_id)
    
    db.session.delete(target)
    db.session.commit()
    
    log_activity('target_delete', f'Hedef silindi: ID {target_id}')
    
    return jsonify({'message': 'Hedef silindi'}), 200

@api.route('/targets/representative/<int:representative_id>', methods=['GET'])
@admin_or_department_manager_required
def get_representative_targets(representative_id):
    """Admin: Belirli bir temsilcinin tüm hedeflerini getir"""
    try:
        # Kullanıcıyı ve erişimi kontrol et
        representative = User.query.get(representative_id)
        if not representative:
            return jsonify({'error': 'Kullanıcı bulunamadı'}), 404
        if not is_user_in_scope(representative_id):
            return jsonify({'error': 'Bu kullanıcıya erişim yetkiniz yok'}), 403
        
        # Temsilcinin tüm hedeflerini getir
        targets = Target.query.filter_by(user_id=representative_id).order_by(Target.year.desc(), Target.month.desc()).all()
        
        targets_data = []
        for target in targets:
            # O ay için satış toplamını hesapla
            sales_total = db.session.query(func.sum(Sales.net_price)).filter(
                Sales.representative_id == representative_id,
                func.extract('year', Sales.date) == target.year,
                func.extract('month', Sales.date) == target.month
            ).scalar() or 0
            
            # O ay için iade toplamını hesapla
            returns_total = db.session.query(func.sum(Returns.net_price)).filter(
                Returns.representative_id == representative_id,
                func.extract('year', Returns.date) == target.year,
                func.extract('month', Returns.date) == target.month
            ).scalar() or 0
            
            net_sales = sales_total - returns_total
            completion_rate = (net_sales / target.target_amount * 100) if target.target_amount > 0 else 0
            
            targets_data.append({
                'id': target.id,
                'year': target.year,
                'month': target.month,
                'target_amount': target.target_amount,
                'total_sales': sales_total,
                'total_returns': returns_total,
                'net_sales': net_sales,
                'completion_rate': completion_rate,
                'created_at': target.created_at.isoformat(),
                'updated_at': target.updated_at.isoformat()
            })
        
        return jsonify({
            'success': True,
            'representative_name': representative.get_full_name(),
            'targets': targets_data
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Satış verileri
@api.route('/sales', methods=['GET'])
@login_required
def get_sales():
    """Satış verilerini getir"""
    # Filtreleme parametreleri
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    representative_id = request.args.get('representative_id', type=int)
    
    query = Sales.query
    
    # Erişim kapsamı uygula
    scoped_ids = get_scoped_user_ids()
    if representative_id:
        # İstenen kullanıcı erişim kapsamındaysa o kullanıcıya göre filtrele
        if not is_user_in_scope(representative_id):
            return jsonify({'error': 'Bu kullanıcıya erişim yetkiniz yok'}), 403
        query = query.filter_by(representative_id=representative_id)
    elif scoped_ids is None:
        # Admin: sınırsız
        pass
    else:
        # DM veya normal kullanıcı: kapsam içi kullanıcılarla sınırla
        query = query.filter(Sales.representative_id.in_(scoped_ids))
    
    # Tarih filtreleme
    if start_date:
        query = query.filter(Sales.date >= start_date)
    if end_date:
        query = query.filter(Sales.date <= end_date)
    
    sales = query.all()
    
    return jsonify({
        'sales': [{
            'id': sale.id,
            'representative_id': sale.representative_id,
            'date': sale.date.isoformat(),
            'product_group': sale.product_group,
            'brand': sale.brand,
            'product_name': sale.product_name,
            'quantity': sale.quantity,
            'unit_price': sale.unit_price,
            'total_price': sale.total_price,
            'net_price': sale.net_price,
            'created_at': sale.created_at.isoformat()
        } for sale in sales]
    }), 200

@api.route('/sales', methods=['POST'])
@permission_required('sales', 'edit')
def create_sale():
    """Temsilci: Satış kaydı oluştur"""
    data = request.get_json()
    
    required_fields = ['date', 'product_group', 'brand', 'product_name', 'quantity', 'unit_price']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'{field} alanı gerekli'}), 400
    
    # Tarih formatını kontrol et
    try:
        sale_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Geçersiz tarih formatı. YYYY-MM-DD kullanın'}), 400
    
    # Toplam fiyat hesapla
    total_price = data['quantity'] * data['unit_price']
    net_price = total_price  # Vergi hesaplaması eklenebilir
    
    sale = Sales(
        representative_id=current_user.id,
        date=sale_date,
        product_group=data['product_group'],
        brand=data['brand'],
        product_name=data['product_name'],
        quantity=data['quantity'],
        unit_price=data['unit_price'],
        total_price=total_price,
        net_price=net_price
    )
    
    db.session.add(sale)
    db.session.commit()
    
    log_activity('sale_create', f'Satış kaydı oluşturuldu: {data["product_name"]}')
    
    return jsonify({'message': 'Satış kaydı oluşturuldu', 'id': sale.id}), 201

# İade verileri
@api.route('/returns', methods=['GET'])
@login_required
def get_returns():
    """İade verilerini getir"""
    # Filtreleme parametreleri
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    representative_id = request.args.get('representative_id', type=int)
    
    query = Returns.query
    
    # Erişim kapsamı uygula
    scoped_ids = get_scoped_user_ids()
    if representative_id:
        if not is_user_in_scope(representative_id):
            return jsonify({'error': 'Bu kullanıcıya erişim yetkiniz yok'}), 403
        query = query.filter_by(representative_id=representative_id)
    elif scoped_ids is None:
        pass
    else:
        query = query.filter(Returns.representative_id.in_(scoped_ids))
    
    # Tarih filtreleme
    if start_date:
        query = query.filter(Returns.date >= start_date)
    if end_date:
        query = query.filter(Returns.date <= end_date)
    
    returns = query.all()
    
    return jsonify({
        'returns': [{
            'id': ret.id,
            'representative_id': ret.representative_id,
            'date': ret.date.isoformat(),
            'product_group': ret.product_group,
            'brand': ret.brand,
            'product_name': ret.product_name,
            'quantity': ret.quantity,
            'unit_price': ret.unit_price,
            'total_price': ret.total_price,
            'net_price': ret.net_price,
            'return_reason': ret.return_reason,
            'created_at': ret.created_at.isoformat()
        } for ret in returns]
    }), 200

@api.route('/returns', methods=['POST'])
@permission_required('sales', 'edit')
def create_return():
    """Temsilci: İade kaydı oluştur"""
    data = request.get_json()
    
    required_fields = ['date', 'product_group', 'brand', 'product_name', 'quantity', 'unit_price']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'{field} alanı gerekli'}), 400
    
    # Tarih formatını kontrol et
    try:
        return_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Geçersiz tarih formatı. YYYY-MM-DD kullanın'}), 400
    
    # Toplam fiyat hesapla
    total_price = data['quantity'] * data['unit_price']
    net_price = total_price
    
    ret = Returns(
        representative_id=current_user.id,
        date=return_date,
        product_group=data['product_group'],
        brand=data['brand'],
        product_name=data['product_name'],
        quantity=data['quantity'],
        unit_price=data['unit_price'],
        total_price=total_price,
        net_price=net_price,
        return_reason=data.get('return_reason')
    )
    
    db.session.add(ret)
    db.session.commit()
    
    log_activity('return_create', f'İade kaydı oluşturuldu: {data["product_name"]}')
    
    return jsonify({'message': 'İade kaydı oluşturuldu', 'id': ret.id}), 201

# Raporlama
@api.route('/reports/summary', methods=['GET'])
@login_required
def get_summary_report():
    """Özet rapor"""
    # Filtreleme parametreleri
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    representative_id = request.args.get('representative_id', type=int)
    
    # Satış sorgusu
    sales_query = Sales.query
    returns_query = Returns.query
    
    # Erişim kapsamı
    scoped_ids = get_scoped_user_ids()
    if representative_id:
        if not is_user_in_scope(representative_id):
            return jsonify({'error': 'Bu kullanıcıya erişim yetkiniz yok'}), 403
        sales_query = sales_query.filter_by(representative_id=representative_id)
        returns_query = returns_query.filter_by(representative_id=representative_id)
    elif scoped_ids is None:
        # Admin: tüm kullanıcılar
        pass
    else:
        sales_query = sales_query.filter(Sales.representative_id.in_(scoped_ids))
        returns_query = returns_query.filter(Returns.representative_id.in_(scoped_ids))
    
    # Tarih filtreleme
    if start_date:
        sales_query = sales_query.filter(Sales.date >= start_date)
        returns_query = returns_query.filter(Returns.date >= start_date)
    if end_date:
        sales_query = sales_query.filter(Sales.date <= end_date)
        returns_query = returns_query.filter(Returns.date <= end_date)
    
    # Toplam değerler
    total_sales = sales_query.with_entities(func.sum(Sales.net_price)).scalar() or 0
    total_returns = returns_query.with_entities(func.sum(Returns.net_price)).scalar() or 0
    net_sales = total_sales - total_returns
    
    # Hedef bilgisi
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    if representative_id:
        # Admin belirli bir temsilcinin hedefini görür
        target = Target.query.filter_by(
            user_id=representative_id,
            year=current_year,
            month=current_month
        ).first()
        target_amount = target.target_amount if target else 0
    elif scoped_ids is None:
        # Admin: tüm hedeflerin toplamı
        target_amount = db.session.query(func.sum(Target.target_amount)).filter_by(
            year=current_year,
            month=current_month
        ).scalar() or 0
    else:
        # DM veya kullanıcı: kapsam içindeki hedefler
        target_amount = db.session.query(func.sum(Target.target_amount)).filter(
            Target.user_id.in_(scoped_ids),
            Target.year == current_year,
            Target.month == current_month
        ).scalar() or 0
    
    target_completion = (net_sales / target_amount * 100) if target_amount > 0 else 0
    
    return jsonify({
        'total_sales': total_sales,
        'total_returns': total_returns,
        'net_sales': net_sales,
        'return_rate': (total_returns / total_sales * 100) if total_sales > 0 else 0,
        'target_amount': target_amount,
        'target_completion': target_completion,
        'period': {
            'start_date': start_date,
            'end_date': end_date
        }
    }), 200

@api.route('/reports/representatives', methods=['GET'])
@admin_or_department_manager_required
def get_representatives_report():
    """Admin/Departman Yöneticisi: kapsamındaki kullanıcıların performans raporu"""
    # Filtreleme parametreleri
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    # Varsayılan: içinde bulunulan ay
    if not start_date and not end_date:
        today = datetime.now().date()
        first = today.replace(day=1)
        # next month first day then minus one day
        if first.month == 12:
            next_first = first.replace(year=first.year+1, month=1, day=1)
        else:
            next_first = first.replace(month=first.month+1, day=1)
        last = next_first - timedelta(days=1)
        start_date = first.isoformat()
        end_date = last.isoformat()
    
    # Erişim kapsamındaki kullanıcıları getir
    scoped_ids = get_scoped_user_ids()
    users_query = User.query.filter(User.is_active == True)
    if scoped_ids is not None:
        users_query = users_query.filter(User.id.in_(scoped_ids))
    
    # Sadece satış departmanı kullanıcılarını getir
    satis_department = Department.query.filter(
        Department.name.ilike('%satış%')
    ).first()
    
    if satis_department:
        users_query = users_query.filter(User.department_id == satis_department.id)
    
    representatives = users_query.all()
    
    report_data = []
    
    for rep in representatives:
        # Satış sorgusu
        sales_query = Sales.query.filter_by(representative_id=rep.id)
        returns_query = Returns.query.filter_by(representative_id=rep.id)
        
        # Tarih filtreleme
        if start_date:
            sales_query = sales_query.filter(Sales.date >= start_date)
            returns_query = returns_query.filter(Returns.date >= start_date)
        if end_date:
            sales_query = sales_query.filter(Sales.date <= end_date)
            returns_query = returns_query.filter(Returns.date <= end_date)
        
        # Toplam değerler
        total_sales = sales_query.with_entities(func.sum(Sales.net_price)).scalar() or 0
        total_returns = returns_query.with_entities(func.sum(Returns.net_price)).scalar() or 0
        net_sales = total_sales - total_returns
        
        # Hedef bilgisi
        current_year = datetime.now().year
        current_month = datetime.now().month
        target = Target.query.filter_by(
            user_id=rep.id,
            year=current_year,
            month=current_month
        ).first()
        
        target_amount = target.target_amount if target else 0
        target_completion = (net_sales / target_amount * 100) if target_amount > 0 else 0
        
        report_data.append({
            'representative_id': rep.id,
            'representative_name': rep.get_full_name() or "Bilinmeyen Temsilci",
            'representative_code': rep.representative_code or "",
            'total_sales': total_sales,
            'total_returns': total_returns,
            'net_sales': net_sales,
            'return_rate': (total_returns / total_sales * 100) if total_sales > 0 else 0,
            'target_amount': target_amount,
            'target_completion': target_completion
        })
    
    return jsonify({'representatives': report_data}), 200



# Aktivite logları
@api.route('/activity-logs', methods=['GET'])
@admin_required
def get_activity_logs():
    """Admin: Aktivite loglarını getir"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'logs': [{
            'id': log.id,
            'user_id': log.user_id,
            'user_name': log.user.get_full_name() if log.user else "Bilinmeyen Kullanıcı",
            'action': log.action,
            'description': log.description,
            'ip_address': log.ip_address,
            'created_at': log.created_at.isoformat()
        } for log in logs.items],
        'total': logs.total,
        'pages': logs.pages,
        'current_page': page
    }), 200 

# Uyumsoft API entegrasyonu için yeni endpoint'ler
@api.route('/uyumsoft/sync', methods=['POST'])
@admin_required
def sync_uyumsoft_data():
    """
    Uyumsoft verilerini sisteme senkronize et
    """
    try:
        from uyumsoft_api import UyumsoftAPI
        from datetime import datetime, timedelta
        
        # Request'ten parametreleri al
        data = request.get_json() or {}
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        # Tarih belirtilmemişse son 30 günü al
        if not end_date:
            end_date = datetime.now().strftime("%Y-%m-%d")
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
        
        # Uyumsoft API konfigürasyonu (config.py'den al)
        from config import Config
        uyumsoft_config = {
            "base_url": Config.UYUMSOFT_API_URL,
            "username": Config.UYUMSOFT_USERNAME,
            "password": Config.UYUMSOFT_PASSWORD,
            "company_id": Config.UYUMSOFT_COMPANY_ID
        }
        
        # API bağlantısını oluştur
        uyumsoft_api = UyumsoftAPI(**uyumsoft_config)
        
        # Satış verilerini çek ve dönüştür
        sales_data = uyumsoft_api.get_sales_data(start_date, end_date)
        transformed_sales = uyumsoft_api.transform_sales_data(sales_data)
        
        # İade verilerini çek ve dönüştür
        returns_data = uyumsoft_api.get_returns_data(start_date, end_date)
        transformed_returns = uyumsoft_api.transform_returns_data(returns_data)
        
        # Verileri sisteme kaydet
        sales_count = 0
        returns_count = 0
        
        for sale_data in transformed_sales:
            try:
                # Tarih formatını düzelt
                if isinstance(sale_data.get('date'), str):
                    sale_data['date'] = datetime.strptime(sale_data['date'], '%Y-%m-%d').date()
                
                sale = Sales(**sale_data)
                db.session.add(sale)
                sales_count += 1
            except Exception as e:
                print(f"Satış verisi kaydetme hatası: {e}")
                continue
        
        for return_data in transformed_returns:
            try:
                # Tarih formatını düzelt
                if isinstance(return_data.get('date'), str):
                    return_data['date'] = datetime.strptime(return_data['date'], '%Y-%m-%d').date()
                
                ret = Returns(**return_data)
                db.session.add(ret)
                returns_count += 1
            except Exception as e:
                print(f"İade verisi kaydetme hatası: {e}")
                continue
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{sales_count} satış ve {returns_count} iade verisi senkronize edildi',
            'sales_count': sales_count,
            'returns_count': returns_count,
            'start_date': start_date,
            'end_date': end_date
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/uyumsoft/test-connection', methods=['POST'])
@admin_required
def test_uyumsoft_connection():
    """
    Uyumsoft API bağlantısını test et
    """
    try:
        from uyumsoft_api import UyumsoftAPI
        from config import Config
        
        uyumsoft_config = {
            "base_url": Config.UYUMSOFT_API_URL,
            "username": Config.UYUMSOFT_USERNAME,
            "password": Config.UYUMSOFT_PASSWORD,
            "company_id": Config.UYUMSOFT_COMPANY_ID
        }
        
        uyumsoft_api = UyumsoftAPI(**uyumsoft_config)
        
        if uyumsoft_api.authenticate():
            return jsonify({
                'success': True,
                'message': 'Uyumsoft API bağlantısı başarılı'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Uyumsoft API bağlantısı başarısız'
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/uyumsoft/sales', methods=['GET'])
@admin_required
def get_uyumsoft_sales():
    """
    Uyumsoft'tan satış verilerini çek (kaydetmeden)
    """
    try:
        from uyumsoft_api import UyumsoftAPI
        from config import Config
        
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        uyumsoft_config = {
            "base_url": Config.UYUMSOFT_API_URL,
            "username": Config.UYUMSOFT_USERNAME,
            "password": Config.UYUMSOFT_PASSWORD,
            "company_id": Config.UYUMSOFT_COMPANY_ID
        }
        
        uyumsoft_api = UyumsoftAPI(**uyumsoft_config)
        sales_data = uyumsoft_api.get_sales_data(start_date, end_date)
        transformed_sales = uyumsoft_api.transform_sales_data(sales_data)
        
        return jsonify({
            'success': True,
            'data': transformed_sales,
            'count': len(transformed_sales)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/uyumsoft/returns', methods=['GET'])
@admin_required
def get_uyumsoft_returns():
    """
    Uyumsoft'tan iade verilerini çek (kaydetmeden)
    """
    try:
        from uyumsoft_api import UyumsoftAPI
        from config import Config
        
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        uyumsoft_config = {
            "base_url": Config.UYUMSOFT_API_URL,
            "username": Config.UYUMSOFT_USERNAME,
            "password": Config.UYUMSOFT_PASSWORD,
            "company_id": Config.UYUMSOFT_COMPANY_ID
        }
        
        uyumsoft_api = UyumsoftAPI(**uyumsoft_config)
        returns_data = uyumsoft_api.get_returns_data(start_date, end_date)
        transformed_returns = uyumsoft_api.transform_returns_data(returns_data)
        
        return jsonify({
            'success': True,
            'data': transformed_returns,
            'count': len(transformed_returns)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500 

@api.route('/sales/recent', methods=['GET'])
@login_required
def get_recent_sales():
    """Son satışları getir (TARIH alanından)"""
    limit = request.args.get('limit', 5, type=int)
    
    query = Sales.query
    
    # Erişim kapsamı
    scoped_ids = get_scoped_user_ids()
    # DM ana sayfada tüm satışları görebilsin (admin gibi)
    if current_user.is_department_manager():
        scoped_ids = None
    if scoped_ids is None:
        pass
    else:
        query = query.filter(Sales.representative_id.in_(scoped_ids))
    
    # Son satışları getir (TARIH alanına göre sırala)
    recent_sales = query.order_by(Sales.date.desc()).limit(limit).all()
    
    return jsonify({
        'sales': [{
            'id': sale.id,
            'representative_id': sale.representative_id,
            'representative_name': sale.representative.get_full_name() if sale.representative else 'Bilinmeyen Temsilci',
            'date': sale.date.isoformat(),
            'date_formatted': sale.original_date or (sale.date.strftime('%d.%m.%Y') if sale.date else 'Tarih Yok'),
            'product_group': sale.original_product_group or sale.product_group,
            'brand': sale.brand,
            'product_name': sale.product_name,
            'quantity': sale.quantity,
            'original_quantity': sale.original_quantity or str(sale.quantity),
            'unit_price': sale.unit_price,
            'total_price': sale.total_price,
            'net_price': sale.net_price,
            'customer_name': sale.customer_name or 'Bilinmeyen Müşteri',
            'customer_code': sale.customer_code or 'Kod Yok',
            'created_at': sale.created_at.isoformat()
        } for sale in recent_sales]
    }), 200

@api.route('/returns/recent', methods=['GET'])
@login_required
def get_recent_returns():
    """Son iadeleri getir (TARIH alanından)"""
    limit = request.args.get('limit', 5, type=int)
    
    query = Returns.query
    
    scoped_ids = get_scoped_user_ids()
    # DM ana sayfada tüm iadeleri görebilsin (admin gibi)
    if current_user.is_department_manager():
        scoped_ids = None
    if scoped_ids is None:
        pass
    else:
        query = query.filter(Returns.representative_id.in_(scoped_ids))
    
    # Son iadeleri getir (TARIH alanına göre sırala)
    recent_returns = query.order_by(Returns.date.desc()).limit(limit).all()
    
    return jsonify({
        'returns': [{
            'id': ret.id,
            'representative_id': ret.representative_id,
            'representative_name': ret.representative.get_full_name() if ret.representative else 'Bilinmeyen Temsilci',
            'date': ret.date.isoformat(),
            'date_formatted': ret.original_date or (ret.date.strftime('%d.%m.%Y') if ret.date else 'Tarih Yok'),
            'product_group': ret.original_product_group or ret.product_group,
            'brand': ret.brand,
            'product_name': ret.product_name,
            'quantity': ret.quantity,
            'original_quantity': ret.original_quantity or str(ret.quantity),
            'unit_price': ret.unit_price,
            'total_price': ret.total_price,
            'net_price': ret.net_price,
            'return_reason': ret.return_reason,
            'customer_name': ret.customer_name or 'Bilinmeyen Müşteri',
            'customer_code': ret.customer_code or 'Kod Yok',
            'created_at': ret.created_at.isoformat()
        } for ret in recent_returns]
    }), 200 

@api.route('/upload-sales-excel', methods=['POST'])
@admin_required
def upload_sales_excel():
    """Excel dosyasından satış verilerini yükle"""
    print("=== Excel Upload Başladı ===")
    print(f"Request files: {list(request.files.keys())}")
    print(f"Request form: {list(request.form.keys())}")
    
    try:
        if 'file' not in request.files:
            print("❌ 'file' key bulunamadı")
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        file = request.files['file']
        print(f"✅ Dosya alındı: {file.filename}")
        
        if file.filename == '':
            print("❌ Dosya adı boş")
            return jsonify({'error': 'Dosya seçilmedi'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            print(f"❌ Geçersiz dosya formatı: {file.filename}")
            return jsonify({'error': 'Sadece Excel dosyaları (.xlsx, .xls) kabul edilir'}), 400
        
        # Excel dosyasını oku (pandas olmadan basit okuma)
        print("📖 Excel dosyası okunuyor...")
        try:
            # Basit Excel okuma - sadece ilk satırı oku
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            
            # Sadece ilk satırı oku (header)
            headers = [cell.value for cell in ws[1]]
            print(f"✅ Excel okundu: {len(headers)} kolon")
            print(f"📊 Excel kolonları: {headers}")
            
            # Gerekli kolonları kontrol et
            required_columns = ['SATISTEMSILCISI', 'TOPLAMNETFIYAT']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                print(f"❌ Eksik kolonlar: {missing_columns}")
                return jsonify({
                    'error': f'Eksik kolonlar: {", ".join(missing_columns)}. Mevcut kolonlar: {", ".join(headers)}'
                }), 400
            
            # Geçici olarak Excel import devre dışı
            return jsonify({
                'message': 'Excel import geçici olarak devre dışı. Pandas paketi eklenene kadar bekleyin.',
                'headers': headers
            }), 200
            
        except Exception as e:
            print(f"❌ Excel okuma hatası: {e}")
            return jsonify({'error': f'Excel okuma hatası: {str(e)}'}), 400
            try:
                # Temsilciyi bul veya oluştur
                representative_name = row['SATISTEMSILCISI']
                
                # Önce tam isim eşleşmesi dene
                representative = User.query.filter(
                    User.first_name + ' ' + User.last_name == representative_name
                ).first()
                
                # Eğer bulunamazsa, username formatında dene
                if not representative:
                    # Excel'deki ismi username formatına çevir
                    username_format = representative_name.lower().replace(' ', '.').replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
                    representative = User.query.filter_by(username=username_format).first()
                
                # Hala bulunamazsa, mevcut username'leri kontrol et
                if not representative:
                    for existing_rep in User.query.filter_by(role=UserRole.REPRESENTATIVE).all():
                        # Excel'deki ismi normalize et
                        excel_name_normalized = representative_name.lower().replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
                        existing_name_normalized = existing_rep.username.lower().replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c')
                        
                        if excel_name_normalized == existing_name_normalized:
                            representative = existing_rep
                            break
                
                if not representative:
                    # Temsilci yoksa oluştur
                    name_parts = representative_name.split()
                    if len(name_parts) >= 2:
                        first_name = name_parts[0]
                        last_name = ' '.join(name_parts[1:])
                    else:
                        first_name = representative_name
                        last_name = ''
                    
                    # Benzersiz temsilci kodu oluştur
                    import random
                    import string
                    
                    # Mevcut kodları kontrol et
                    existing_codes = [user.representative_code for user in User.query.filter_by(role=UserRole.REPRESENTATIVE).all() if user.representative_code]
                    
                    # Benzersiz kod oluştur
                    while True:
                        # T + 3 haneli rastgele sayı
                        code = f"T{random.randint(100, 999)}"
                        if code not in existing_codes:
                            break
                    
                    representative = User(
                        username=representative_name.lower().replace(' ', '.'),
                        email=f"{representative_name.lower().replace(' ', '.')}@company.com",
                        password_hash=generate_password_hash('123456'),
                        role=UserRole.REPRESENTATIVE,
                        first_name=first_name,
                        last_name=last_name,
                        representative_code=code,
                        is_active=True
                    )
                    
                    db.session.add(representative)
                    db.session.flush()  # ID'yi almak için flush
                    created_representatives.append(representative_name)
                    print(f"✅ Yeni temsilci oluşturuldu: {representative_name} (Kod: {code})")
                
                # Satış verisini oluştur - JSON'daki orijinal alanları kullan
                sale_data = {
                    'representative_id': representative.id,
                    'date': pd.to_datetime(row.get('TARIH', datetime.now())).date(),
                    'product_group': row.get('URUN_ANA_GRUP', 'Bilinmeyen') if pd.notna(row.get('URUN_ANA_GRUP')) else 'Bilinmeyen',
                    'brand': row.get('MARKA', 'Bilinmeyen') if pd.notna(row.get('MARKA')) else 'Bilinmeyen',
                    'product_name': row.get('STOKADı', 'Bilinmeyen Ürün') if pd.notna(row.get('STOKADı')) else 'Bilinmeyen Ürün',
                    'quantity': int(row.get('ADET', 1)) if pd.notna(row.get('ADET')) else 1,
                    'unit_price': float(row.get('BIRIMFIYAT', 0)) if pd.notna(row.get('BIRIMFIYAT')) else 0,
                    'total_price': float(row['TOPLAMNETFIYAT']) if pd.notna(row['TOPLAMNETFIYAT']) else 0,
                    'net_price': float(row['TOPLAMNETFIYAT']) if pd.notna(row['TOPLAMNETFIYAT']) else 0,
                    'customer_name': row.get('CARIADI', 'Bilinmeyen Müşteri') if pd.notna(row.get('CARIADI')) else 'Bilinmeyen Müşteri',
                    'customer_code': row.get('CARIKODU', 'Kod Yok') if pd.notna(row.get('CARIKODU')) else 'Kod Yok',
                    # Orijinal alanları kaydet
                    'original_quantity': str(row.get('ADET', '')) if pd.notna(row.get('ADET')) else '',
                    'original_date': str(row.get('TARIH', '')) if pd.notna(row.get('TARIH')) else '',
                    'original_product_group': str(row.get('URUN_ANA_GRUP', '')) if pd.notna(row.get('URUN_ANA_GRUP')) else ''
                }
                
                sale = Sales(**sale_data)
                db.session.add(sale)
                sales_count += 1
                print(f"✅ Satış kaydedildi: {representative_name} - {sale_data['total_price']:,.2f} TL")
                
            except Exception as e:
                print(f"❌ Satır işleme hatası: {e}")
                continue
        
        # Tüm değişiklikleri kaydet
        db.session.commit()
        print(f"✅ Toplam {sales_count} satış kaydedildi")
        
        return jsonify({
            'success': True,
            'message': f'{sales_count} satış kaydedildi',
            'sales_count': sales_count,
            'created_representatives': created_representatives
        })
        
    except Exception as e:
        print(f"❌ Genel hata: {e}")
        db.session.rollback()
        return jsonify({'error': f'Veritabanı hatası: {str(e)}'}), 500

@api.route('/sales/representatives', methods=['GET'])
@login_required
def get_representatives_sales():
    """Kapsamdaki kullanıcıların satış verilerini getir (filtreleme destekli)."""
    try:
        # Filtre parametrelerini al
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        status = request.args.get('status', type=str)
        
        # Erişim kapsamındaki kullanıcılar
        scoped_ids = get_scoped_user_ids()
        query_users = User.query.filter(User.is_active == True)
        if scoped_ids is not None:
            query_users = query_users.filter(User.id.in_(scoped_ids))
        representatives = query_users.all()
        
        result = []
        for rep in representatives:
            # Temsilcinin satışlarını hesapla (filtreleme varsa o dönem için, yoksa bu ay için)
            if year and month:
                target_year = year
                target_month = month
            else:
                target_year = datetime.now().year
                target_month = datetime.now().month
            
            sales_query = Sales.query.filter_by(representative_id=rep.id)
            sales_query = sales_query.filter(
                func.extract('year', Sales.date) == target_year,
                func.extract('month', Sales.date) == target_month
            )
            total_sales = sales_query.with_entities(func.sum(Sales.net_price)).scalar() or 0
            
            # Temsilcinin iadelerini hesapla
            returns_query = Returns.query.filter_by(representative_id=rep.id)
            returns_query = returns_query.filter(
                func.extract('year', Returns.date) == target_year,
                func.extract('month', Returns.date) == target_month
            )
            total_returns = returns_query.with_entities(func.sum(Returns.net_price)).scalar() or 0
            
            # Net satış hesapla
            net_sales = total_sales - total_returns
            
            # Hedefi getir (DM için departman kapsamı otomatik sağlanır)
            target = Target.query.filter_by(user_id=rep.id, year=target_year, month=target_month).first()
            
            target_amount = target.target_amount if target else 0
            completion_rate = (net_sales / target_amount * 100) if target_amount > 0 else 0
            
            # Durum filtresini uygula
            if status:
                if status == 'completed' and completion_rate < 100:
                    continue
                elif status == 'good' and (completion_rate < 80 or completion_rate >= 100):
                    continue
                elif status == 'low' and completion_rate >= 80:
                    continue
            
            # Ürün gruplarına göre dağılım
            product_groups = db.session.query(
                Sales.product_group,
                func.sum(Sales.net_price).label('total')
            ).filter_by(representative_id=rep.id).group_by(Sales.product_group).all()
            
            # Markalara göre dağılım
            brands = db.session.query(
                Sales.brand,
                func.sum(Sales.net_price).label('total')
            ).filter_by(representative_id=rep.id).group_by(Sales.brand).all()
            
            result.append({
                'representative_id': rep.id,
                'representative_name': rep.get_full_name() or "Bilinmeyen Temsilci",
                'representative_code': rep.representative_code or "",
                'total_sales': total_sales,
                'total_returns': total_returns,
                'net_sales': net_sales,
                'target_amount': target_amount,
                'completion_rate': completion_rate,
                'product_groups': [{'name': pg.product_group or "Bilinmeyen", 'total': pg.total} for pg in product_groups],
                'brands': [{'name': b.brand or "Bilinmeyen", 'total': b.total} for b in brands]
            })
        
        return jsonify({
            'success': True,
            'representatives': result,
            'filters': {
                'year': year,
                'month': month,
                'status': status
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/sales/charts-data', methods=['GET'])
@login_required
def get_sales_charts_data():
    """Grafikler için satış verilerini getir"""
    try:
        # Rol bazlı filtreleme: Admin ve DM tüm veriyi görür
        query_filter = {}
        if not (current_user.is_admin() or current_user.is_department_manager()):
            query_filter['representative_id'] = current_user.id
        
        # Temsilci bazlı satış verileri - SQLite uyumlu
        rep_query = db.session.query(
            User.first_name,
            User.last_name,
            func.sum(Sales.net_price).label('total_sales')
        ).join(Sales, User.id == Sales.representative_id)
        
        if query_filter:
            rep_query = rep_query.filter(Sales.representative_id == query_filter['representative_id'])
        
        rep_sales = rep_query.group_by(User.id).all()
        
        # Marka bazlı satış verileri
        brand_query = db.session.query(
            Sales.brand,
            func.sum(Sales.net_price).label('total_sales')
        )
        if query_filter:
            brand_query = brand_query.filter(Sales.representative_id == query_filter['representative_id'])
        brand_sales = brand_query.group_by(Sales.brand).all()
        
        # Ürün grubu bazlı satış verileri
        product_query = db.session.query(
            Sales.product_group,
            func.sum(Sales.net_price).label('total_sales')
        )
        if query_filter:
            product_query = product_query.filter(Sales.representative_id == query_filter['representative_id'])
        product_sales = product_query.group_by(Sales.product_group).all()

        # Marka ve Ürün Grubu kombinasyonu
        brand_product_query = db.session.query(
            Sales.brand,
            Sales.product_group,
            func.sum(Sales.net_price).label('total_sales')
        )
        if query_filter:
            brand_product_query = brand_product_query.filter(Sales.representative_id == query_filter['representative_id'])
        brand_product_sales = brand_product_query.group_by(Sales.brand, Sales.product_group).all()
        
        # Aylık satış trendi (son 12 ay) - SQLite uyumlu
        monthly_query = db.session.query(
            func.strftime('%Y-%m', Sales.date).label('month'),
            func.sum(Sales.net_price).label('total_sales')
        )
        if query_filter:
            monthly_query = monthly_query.filter(Sales.representative_id == query_filter['representative_id'])
        monthly_sales = monthly_query.group_by(func.strftime('%Y-%m', Sales.date)).order_by('month').limit(12).all()

        return jsonify({
            'success': True,
            'representative_sales': [{'name': f"{rep[0] or ''} {rep[1] or ''}".strip() or "Bilinmeyen", 'total': float(rep[2])} for rep in rep_sales],
            'product_sales': [{'name': ps[0] or "Bilinmeyen", 'total': float(ps[1])} for ps in product_sales],
            'brand_sales': [{'name': bs[0] or "Bilinmeyen", 'total': float(bs[1])} for bs in brand_sales],
            'brand_product_sales': [{'brand': bp[0] or "Bilinmeyen", 'product_group': bp[1] or "Bilinmeyen", 'total': float(bp[2])} for bp in brand_product_sales],
            'monthly_trend': [{'month': ms[0], 'total': float(ms[1])} for ms in monthly_sales]
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/users/representatives', methods=['GET'])
@admin_or_department_manager_required
def get_representatives():
    """Departman yöneticisi: kendi departmanındaki kullanıcıları, Admin: tüm kullanıcıları getirir.
    Eski 'temsilci' kavramı yerine departman kullanıcıları listelenir.
    """
    try:
        scoped_ids = get_scoped_user_ids()
        query = User.query.filter(User.is_active == True)
        if scoped_ids is not None:
            query = query.filter(User.id.in_(scoped_ids))
        users = query.all()
        return jsonify({
            'success': True,
            'representatives': [{
                'id': u.id,
                'name': u.get_full_name(),
                'username': u.username or "",
                'email': u.email or "",
                'representative_code': u.representative_code or "",
                'is_active': u.is_active
            } for u in users]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/representatives', methods=['GET'])
@login_required
def get_representatives_for_planning():
    """Planlama/atama için kullanıcı listesi.
    Güncel ihtiyaç: Herkes herkese görev atayabildiği için tüm kullanıcıları döner."""
    try:
        users = User.query.filter(User.is_active == True).all()
        return jsonify({
            'success': True,
            'representatives': [{
                'id': u.id,
                'first_name': u.first_name,
                'last_name': u.last_name,
                'representative_code': u.representative_code,
                'region': u.region,
                'phone': u.phone,
                'department_id': u.department_id
            } for u in users]
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/targets/bulk-create', methods=['POST'])
@admin_required
def bulk_create_targets():
    """Toplu hedef oluştur"""
    try:
        data = request.get_json()
        targets_data = data.get('targets', [])
        
        if not targets_data:
            return jsonify({'error': 'Hedef verileri gerekli'}), 400
        
        created_targets = []
        current_year = datetime.now().year
        current_month = datetime.now().month
        
        for target_data in targets_data:
            try:
                representative_name = target_data.get('representative_name')
                target_amount = target_data.get('target_amount')
                
                if not representative_name or not target_amount:
                    continue
                
                # Temsilciyi bul
                representative = User.query.filter(
                    User.first_name + ' ' + User.last_name == representative_name
                ).first()
                
                if not representative:
                    continue
                
                # Aynı ay için hedef var mı kontrol et
                existing_target = Target.query.filter_by(
                    user_id=representative.id,
                    year=current_year,
                    month=current_month
                ).first()
                
                if existing_target:
                    # Mevcut hedefi güncelle
                    existing_target.target_amount = target_amount
                    existing_target.updated_at = datetime.now()
                else:
                    # Yeni hedef oluştur
                    target = Target(
                        user_id=representative.id,
                        year=current_year,
                        month=current_month,
                        target_amount=target_amount
                    )
                    db.session.add(target)
                
                created_targets.append({
                    'representative_name': representative_name,
                    'target_amount': target_amount
                })
                
            except Exception as e:
                print(f"Hedef oluşturma hatası: {e}")
                continue
        
        db.session.commit()
        
        log_activity('bulk_target_create', f'{len(created_targets)} hedef oluşturuldu/güncellendi')
        
        return jsonify({
            'success': True,
            'message': f'{len(created_targets)} hedef başarıyla oluşturuldu/güncellendi',
            'targets': created_targets
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': f'Hedef oluşturma hatası: {str(e)}'
        }), 500 

@api.route('/sales/representative', methods=['GET'])
@login_required
def get_representative_sales():
    """Giriş yapan kullanıcının (erişim kapsamına göre) satış verilerini getir."""
    try:
        # Kullanıcının satış verisi görüntüleme yetkisi olmalı
        if not current_user.has_permission('sales', 'view'):
            return jsonify({'error': 'Satış verilerini görüntüleme yetkiniz yok'}), 403
        
        # Kullanıcının satış verilerini getir
        sales_query = Sales.query.filter_by(representative_id=current_user.id)
        
        # Toplam satış
        total_sales = sales_query.with_entities(func.sum(Sales.net_price)).scalar() or 0
        
        # Toplam iade
        returns_query = Returns.query.filter_by(representative_id=current_user.id)
        total_returns = returns_query.with_entities(func.sum(Returns.net_price)).scalar() or 0
        
        # Net satış
        net_sales = total_sales - total_returns
        
        # Hedef bilgileri
        current_year = datetime.now().year
        current_month = datetime.now().month
        target = Target.query.filter_by(
            user_id=current_user.id,
            year=current_year,
            month=current_month
        ).first()
        
        target_amount = target.target_amount if target else 0
        target_completion = (net_sales / target_amount * 100) if target_amount > 0 else 0
        
        # Aylık satış verileri (son 6 ay)
        monthly_sales = []
        for i in range(6):
            month_date = datetime.now() - timedelta(days=30*i)
            month_start = month_date.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_sales = sales_query.filter(
                Sales.date >= month_start.date(),
                Sales.date <= month_end.date()
            ).with_entities(func.sum(Sales.net_price)).scalar() or 0
            
            monthly_sales.append({
                'month': month_start.strftime('%B %Y'),
                'total': month_sales
            })
        
        monthly_sales.reverse()  # En eski aydan en yeni aya
        
        # Ürün grupları
        product_groups = db.session.query(
            Sales.product_group,
            func.sum(Sales.net_price).label('total')
        ).filter_by(representative_id=current_user.id).group_by(Sales.product_group).all()
        
        # Marka performansı
        brand_performance = db.session.query(
            Sales.brand,
            func.sum(Sales.net_price).label('total')
        ).filter_by(representative_id=current_user.id).group_by(Sales.brand).order_by(func.sum(Sales.net_price).desc()).limit(5).all()
        
        # En iyi ürünler
        best_products = db.session.query(
            Sales.product_name,
            func.sum(Sales.net_price).label('total')
        ).filter_by(representative_id=current_user.id).group_by(Sales.product_name).order_by(func.sum(Sales.net_price).desc()).limit(5).all()
        
        # Son satışlar
        recent_sales = sales_query.order_by(Sales.date.desc()).limit(10).all()
        
        return jsonify({
            'success': True,
            'total_sales': total_sales,
            'total_returns': total_returns,
            'net_sales': net_sales,
            'target_amount': target_amount,
            'target_completion': target_completion,
            'monthly_sales': monthly_sales,
            'product_groups': [{'group': pg.product_group or "Bilinmeyen", 'total': pg.total} for pg in product_groups],
            'brand_performance': [{'name': bp.brand or "Bilinmeyen", 'total': bp.total} for bp in brand_performance],
            'best_products': [{'name': bp.product_name or "Bilinmeyen", 'total': bp.total} for bp in best_products],
            'recent_sales': [{
                'date': sale.date.isoformat(),
                'product_name': sale.product_name,
                'brand': sale.brand,
                'quantity': sale.quantity,
                'unit_price': sale.unit_price,
                'total_price': sale.total_price,
                'customer_name': sale.customer_name
            } for sale in recent_sales]
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500 

# Planlama endpoint'leri
## Planning endpoints removed

## Planning endpoints removed

## Planning endpoints removed

## Planning endpoints removed

## Planning endpoints removed

## Planning endpoints removed

## Planning endpoints removed

## Planning endpoints removed

@api.route('/plans', methods=['GET'])
@login_required
def get_plans():
    """Aylık hedef planlarını getir"""
    try:
        filter_type = request.args.get('filter', 'all')
        
        query = Target.query
        
        # Admin tüm planları görebilir, temsilci sadece kendininkini
        if not current_user.is_admin():
            query = query.filter_by(user_id=current_user.id)
        
        # Filtreleme
        if filter_type == 'active':
            # Aktif planlar (bu ay ve gelecek aylar)
            current_date = datetime.now()
            query = query.filter(
                db.or_(
                    db.and_(Target.year == current_date.year, Target.month >= current_date.month),
                    Target.year > current_date.year
                )
            )
        elif filter_type == 'completed':
            # Tamamlanan planlar (geçmiş aylar)
            current_date = datetime.now()
            query = query.filter(
                db.or_(
                    db.and_(Target.year == current_date.year, Target.month < current_date.month),
                    Target.year < current_date.year
                )
            )
        
        plans = query.order_by(Target.year.desc(), Target.month.desc()).all()
        
        # Her plan için gerçekleşen satış miktarını hesapla
        plans_data = []
        for plan in plans:
            # O ay için satış toplamını hesapla
            sales_total = db.session.query(func.sum(Sales.total_price)).filter(
                Sales.representative_id == plan.user_id,
                func.extract('year', Sales.date) == plan.year,
                func.extract('month', Sales.date) == plan.month
            ).scalar() or 0
            
            # O ay için iade toplamını hesapla
            returns_total = db.session.query(func.sum(Returns.total_price)).filter(
                Returns.representative_id == plan.user_id,
                func.extract('year', Returns.date) == plan.year,
                func.extract('month', Returns.date) == plan.month
            ).scalar() or 0
            
            actual_amount = sales_total - returns_total
            
            plans_data.append({
                'id': plan.id,
                'plan_month': f"{plan.year}-{plan.month:02d}",
                'year': plan.year,
                'month': plan.month,
                'target_amount': plan.target_amount,
                'actual_amount': actual_amount,
                'region': plan.user.region,
                'representative_id': plan.user_id,
                'representative_name': plan.user.get_full_name(),
                'created_at': plan.created_at.isoformat(),
                'updated_at': plan.updated_at.isoformat()
            })
        
        return jsonify({
            'success': True,
            'plans': plans_data
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/plans/current-month-sales', methods=['GET'])
@login_required
def get_current_month_sales():
    """Bu ay için satış verilerini getir (planı olmayan kullanıcılar için)"""
    try:
        current_month = datetime.now().strftime('%Y-%m')
        year, month_num = map(int, current_month.split('-'))
        
        if current_user.is_admin():
            # Admin için tüm satışları getir
            sales_total = db.session.query(func.sum(Sales.total_price)).filter(
                func.extract('year', Sales.date) == year,
                func.extract('month', Sales.date) == month_num
            ).scalar() or 0
            
            returns_total = db.session.query(func.sum(Returns.total_price)).filter(
                func.extract('year', Returns.date) == year,
                func.extract('month', Returns.date) == month_num
            ).scalar() or 0
        else:
            # Temsilci için kendi satışlarını getir
            sales_total = db.session.query(func.sum(Sales.total_price)).filter(
                Sales.representative_id == current_user.id,
                func.extract('year', Sales.date) == year,
                func.extract('month', Sales.date) == month_num
            ).scalar() or 0
            
            returns_total = db.session.query(func.sum(Returns.total_price)).filter(
                Returns.representative_id == current_user.id,
                func.extract('year', Returns.date) == year,
                func.extract('month', Returns.date) == month_num
            ).scalar() or 0
        
        actual_amount = sales_total - returns_total
        
        return jsonify({
            'success': True,
            'actual_amount': actual_amount,
            'sales_total': sales_total,
            'returns_total': returns_total
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/plans', methods=['POST'])
@login_required
def create_plan():
    """Yeni aylık hedef planı oluştur"""
    try:
        data = request.get_json()
        
        # Gerekli alanları kontrol et
        required_fields = ['planMonth', 'targetAmount', 'region']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'{field} alanı gerekli'
                }), 400
        
        # Plan ayını parse et
        try:
            plan_date = datetime.strptime(data['planMonth'], '%Y-%m')
            year = plan_date.year
            month = plan_date.month
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Geçersiz tarih formatı'
            }), 400
        
        # Temsilci ID'sini belirle
        representative_id = data.get('representative')
        if not representative_id and current_user.is_admin():
            return jsonify({
                'success': False,
                'error': 'Temsilci seçimi gerekli'
            }), 400
        
        if not representative_id:
            representative_id = current_user.id
        
        # Temsilciyi kontrol et
        representative = User.query.get(representative_id)
        if not representative or not representative.is_representative():
            return jsonify({
                'success': False,
                'error': 'Geçersiz temsilci'
            }), 400
        
        # Aynı ay için plan var mı kontrol et
        existing_plan = Target.query.filter_by(
            user_id=representative_id,
            year=year,
            month=month
        ).first()
        
        if existing_plan:
            return jsonify({
                'success': False,
                'error': f'{year}/{month:02d} ayı için zaten plan mevcut'
            }), 400
        
        # Yeni plan oluştur
        new_plan = Target(
            user_id=representative_id,
            year=year,
            month=month,
            target_amount=float(data['targetAmount'])
        )
        
        db.session.add(new_plan)
        db.session.commit()
        
        log_activity('plan_create', f'Aylık hedef planı oluşturuldu: {representative.get_full_name()} - {year}/{month:02d}')
        
        return jsonify({
            'success': True,
            'message': 'Plan başarıyla oluşturuldu'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/plans/<int:plan_id>', methods=['GET'])
@login_required
def get_plan(plan_id):
    """Belirli bir planı getir"""
    try:
        plan = Target.query.get_or_404(plan_id)
        
        # Yetki kontrolü
        if not current_user.is_admin() and plan.user_id != current_user.id:
            return jsonify({
                'success': False,
                'error': 'Bu plana erişim yetkiniz yok'
            }), 403
        
        # Gerçekleşen satış miktarını hesapla
        sales_total = db.session.query(func.sum(Sales.total_price)).filter(
            Sales.representative_id == plan.user_id,
            func.extract('year', Sales.date) == plan.year,
            func.extract('month', Sales.date) == plan.month
        ).scalar() or 0
        
        returns_total = db.session.query(func.sum(Returns.total_price)).filter(
            Returns.representative_id == plan.user_id,
            func.extract('year', Returns.date) == plan.year,
            func.extract('month', Returns.date) == plan.month
        ).scalar() or 0
        
        actual_amount = sales_total - returns_total
        
        return jsonify({
            'success': True,
            'plan': {
                'id': plan.id,
                'plan_month': f"{plan.year}-{plan.month:02d}",
                'year': plan.year,
                'month': plan.month,
                'target_amount': plan.target_amount,
                'actual_amount': actual_amount,
                'region': plan.user.region,
                'representative_id': plan.user_id,
                'representative_name': plan.user.get_full_name(),
                'created_at': plan.created_at.isoformat(),
                'updated_at': plan.updated_at.isoformat()
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/plans/<int:plan_id>', methods=['PUT'])
@login_required
def update_plan(plan_id):
    """Plan güncelle"""
    try:
        plan = Target.query.get_or_404(plan_id)
        
        # Yetki kontrolü
        if not current_user.is_admin() and plan.user_id != current_user.id:
            return jsonify({
                'success': False,
                'error': 'Bu planı güncelleyemezsiniz'
            }), 403
        
        data = request.get_json()
        
        if 'targetAmount' in data:
            plan.target_amount = float(data['targetAmount'])
        
        if 'planMonth' in data:
            try:
                plan_date = datetime.strptime(data['planMonth'], '%Y-%m')
                plan.year = plan_date.year
                plan.month = plan_date.month
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': 'Geçersiz tarih formatı'
                }), 400
        
        db.session.commit()
        
        log_activity('plan_update', f'Plan güncellendi: ID {plan_id}')
        
        return jsonify({
            'success': True,
            'message': 'Plan başarıyla güncellendi'
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@api.route('/plans/<int:plan_id>', methods=['DELETE'])
@login_required
def delete_plan(plan_id):
    """Plan sil"""
    try:
        plan = Target.query.get_or_404(plan_id)
        
        # Yetki kontrolü
        if not current_user.is_admin() and plan.user_id != current_user.id:
            return jsonify({
                'success': False,
                'error': 'Bu planı silemezsiniz'
            }), 403
        
        db.session.delete(plan)
        db.session.commit()
        
        log_activity('plan_delete', f'Plan silindi: ID {plan_id}')
        
        return jsonify({
            'success': True,
            'message': 'Plan başarıyla silindi'
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api.route('/users/<int:user_id>/hard-delete', methods=['DELETE'])
@admin_required
def hard_delete_user(user_id):
    """Admin: Kullanıcıyı kalıcı olarak sil.
    Varsayılan: kayıtlar varsa reassign_to zorunlu. purge=1 gönderilirse tüm ilişkili kayıtlar silinir."""
    try:
        user = User.query.get_or_404(user_id)
        if user.id == current_user.id:
            return jsonify({'success': False, 'error': 'Kendinizi silemezsiniz'}), 400
        if user.role == UserRole.ADMIN:
            admin_count = User.query.filter_by(role=UserRole.ADMIN, is_active=True).count()
            if admin_count <= 1:
                return jsonify({'success': False, 'error': 'Son aktif admin silinemez'}), 400

        reassign_to = request.args.get('reassign_to', type=int)
        purge = request.args.get('purge', default=0, type=int) == 1
        sales_count = Sales.query.filter_by(representative_id=user.id).count()
        returns_count = Returns.query.filter_by(representative_id=user.id).count()
        target_count = Target.query.filter_by(user_id=user.id).count()
        tasks_created_count = Task.query.filter_by(created_by_id=user.id).count()
        task_comments_count = TaskComment.query.filter_by(user_id=user.id).count() if 'TaskComment' in globals() else 0
        needs_reassign = (sales_count + returns_count + target_count + tasks_created_count + task_comments_count) > 0

        target_user = None
        if needs_reassign and not purge:
            if not reassign_to:
                return jsonify({
                    'success': False,
                    'error': 'Kayıtlar mevcut. Önce devretmelisiniz ya da purge=1 kullanın.',
                    'needs_reassign': True,
                    'counts': {
                        'sales': sales_count,
                        'returns': returns_count,
                        'targets': target_count,
                        'tasks_created': tasks_created_count,
                        'task_comments': task_comments_count
                    }
                }), 400
            target_user = User.query.get_or_404(reassign_to)

        # Devir yapılacaklar veya purge
        if purge:
            # Satış/İade/Hedef ve kullanıcıya bağlı görev/yorumları sil
            Sales.query.filter_by(representative_id=user.id).delete(synchronize_session=False)
            Returns.query.filter_by(representative_id=user.id).delete(synchronize_session=False)
            Target.query.filter_by(user_id=user.id).delete(synchronize_session=False)
            # Görevler: kullanıcının oluşturduğu ya da atadığı/atanan olduğu tüm görevler silinsin
            Task.query.filter(
                (Task.created_by_id == user.id) | (Task.assigned_by_id == user.id) | (Task.assigned_to_id == user.id)
            ).delete(synchronize_session=False)
            try:
                TaskComment.query.filter_by(user_id=user.id).delete(synchronize_session=False)
            except Exception:
                pass
        elif target_user:
            Sales.query.filter_by(representative_id=user.id).update({Sales.representative_id: target_user.id})
            Returns.query.filter_by(representative_id=user.id).update({Returns.representative_id: target_user.id})
            Target.query.filter_by(user_id=user.id).update({Target.user_id: target_user.id})
            Task.query.filter_by(created_by_id=user.id).update({Task.created_by_id: target_user.id})
            try:
                TaskComment.query.filter_by(user_id=user.id).update({TaskComment.user_id: target_user.id})
            except Exception:
                pass

        # Planlama kayıtlarını temizle
        Planning.query.filter_by(representative_id=user.id).delete(synchronize_session=False)
        PlanningSnapshot.query.filter_by(representative_id=user.id).delete(synchronize_session=False)

        # Bildirimler ve görev atamaları
        Notification.query.filter_by(to_user_id=user.id).delete(synchronize_session=False)
        if purge:
            Notification.query.filter_by(created_by_id=user.id).delete(synchronize_session=False)
        else:
            Notification.query.filter_by(created_by_id=user.id).update({Notification.created_by_id: None})
        if not purge:
            Task.query.filter_by(assigned_by_id=user.id).update({Task.assigned_by_id: None})
            Task.query.filter_by(assigned_to_id=user.id).update({Task.assigned_to_id: None})

        # Departman yöneticiliğini boşalt
        Department.query.filter_by(manager_id=user.id).update({Department.manager_id: None})

        # Aktivite logları
        ActivityLog.query.filter_by(user_id=user.id).delete(synchronize_session=False)

        db.session.commit()

        # Kullanıcıyı sil
        db.session.delete(user)
        db.session.commit()

        try:
            log = ActivityLog(user_id=current_user.id, action='user_hard_deleted', description=f'Kullanıcı kalıcı silindi: {user_id}')
            db.session.add(log)
            db.session.commit()
        except Exception:
            db.session.rollback()

        return jsonify({'success': True, 'message': 'Kullanıcı kalıcı olarak silindi'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@api.route('/plans/summary', methods=['GET'])
@login_required
def get_planning_summary():
    """Aylık planlama özeti"""
    try:
        month = request.args.get('month')
        if not month:
            month = datetime.now().strftime('%Y-%m')
        
        year, month_num = map(int, month.split('-'))
        
        # Kullanıcının planlarını getir
        if current_user.is_admin():
            # Admin tüm planları görebilir
            plans = Target.query.filter_by(year=year, month=month_num).all()
            user_ids = [plan.user_id for plan in plans]
        else:
            # Temsilci sadece kendi planını görebilir
            plans = Target.query.filter_by(
                user_id=current_user.id,
                year=year,
                month=month_num
            ).all()
            user_ids = [current_user.id]
        
        total_target = sum(plan.target_amount for plan in plans)
        
        # Gerçekleşen satışları hesapla - planı olmayan kullanıcılar için de hesapla
        total_actual = 0
        
        # Planı olan kullanıcılar için satış hesapla
        for plan in plans:
            sales_total = db.session.query(func.sum(Sales.total_price)).filter(
                Sales.representative_id == plan.user_id,
                func.extract('year', Sales.date) == year,
                func.extract('month', Sales.date) == month_num
            ).scalar() or 0
            
            returns_total = db.session.query(func.sum(Returns.total_price)).filter(
                Returns.representative_id == plan.user_id,
                func.extract('year', Returns.date) == year,
                func.extract('month', Returns.date) == month_num
            ).scalar() or 0
            
            total_actual += sales_total - returns_total
        
        # Planı olmayan kullanıcılar için de satış hesapla
        if current_user.is_admin():
            # Admin için tüm temsilcilerin satışlarını hesapla
            all_sales = db.session.query(func.sum(Sales.total_price)).filter(
                func.extract('year', Sales.date) == year,
                func.extract('month', Sales.date) == month_num
            ).scalar() or 0
            
            all_returns = db.session.query(func.sum(Returns.total_price)).filter(
                func.extract('year', Returns.date) == year,
                func.extract('month', Returns.date) == month_num
            ).scalar() or 0
            
            total_actual = all_sales - all_returns
        else:
            # Temsilci için kendi satışlarını hesapla (planı olmasa bile)
            if not plans:  # Eğer plan yoksa
                sales_total = db.session.query(func.sum(Sales.total_price)).filter(
                    Sales.representative_id == current_user.id,
                    func.extract('year', Sales.date) == year,
                    func.extract('month', Sales.date) == month_num
                ).scalar() or 0
                
                returns_total = db.session.query(func.sum(Returns.total_price)).filter(
                    Returns.representative_id == current_user.id,
                    func.extract('year', Returns.date) == year,
                    func.extract('month', Returns.date) == month_num
                ).scalar() or 0
                
                total_actual = sales_total - returns_total
        
        remaining = max(0, total_target - total_actual)
        
        return jsonify({
            'success': True,
            'target': total_target,
            'actual': total_actual,
            'remaining': remaining
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500 


# Kullanıcı Yönetimi API Endpoint'leri
@api.route('/users', methods=['GET'])
@admin_required
def get_all_users():
    """Tüm kullanıcıları getir (varsayılan: sadece aktif kullanıcılar).
    include_inactive=1 gönderilirse pasif kullanıcılar da dahil edilir."""
    try:
        include_inactive = request.args.get('include_inactive', type=int)
        query = User.query
        if not include_inactive:
            query = query.filter(User.is_active == True)
        users = query.all()
        users_data = []
        
        for user in users:
            users_data.append({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'role': user.role.value if user.role else None,
                'department_id': user.department_id,
                'department_role': user.department_role,
                'department_default_role_title': user.department.default_role_title if user.department else None,
                'representative_code': user.representative_code,
                'is_active': user.is_active,
                'created_at': user.created_at.isoformat() if user.created_at else None,
                'last_login': user.last_login.isoformat() if user.last_login else None
            })
        
        return jsonify({
            'success': True,
            'users': users_data
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api.route('/users', methods=['POST'])
@admin_required
def create_user():
    """Yeni kullanıcı oluştur"""
    try:
        data = request.get_json()
        
        # Gerekli alanları kontrol et
        required_fields = ['username', 'password', 'first_name', 'last_name', 'role']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'{field} alanı gereklidir'
                }), 400
        
        # Kullanıcı adı kontrolü
        if User.query.filter_by(username=data['username']).first():
            return jsonify({
                'success': False,
                'error': 'Bu kullanıcı adı zaten kullanılıyor'
            }), 400
        
        # Email kontrolü (eğer email verilmişse)
        if data.get('email') and User.query.filter_by(email=data['email']).first():
            return jsonify({
                'success': False,
                'error': 'Bu email adresi zaten kullanılıyor'
            }), 400
        
        # Kullanıcı kodu: boş string -> None; benzersiz kontrol (opsiyonel)
        rep_code = data.get('representative_code')
        if isinstance(rep_code, str):
            rep_code = rep_code.strip()
        if not rep_code:
            rep_code = None
        else:
            if User.query.filter_by(representative_code=rep_code).first():
                return jsonify({'success': False,'error': 'Bu kullanıcı kodu zaten kullanılıyor'}), 400
        
        # Role string'ini UserRole enum'ına dönüştür
        role_mapping = {
            'ADMIN': UserRole.ADMIN,
            'DEPARTMENT_MANAGER': UserRole.DEPARTMENT_MANAGER,
            'USER': UserRole.USER
        }
        
        if data['role'] not in role_mapping:
            return jsonify({
                'success': False,
                'error': f'Geçersiz rol: {data["role"]}'
            }), 400
        
        # Yeni kullanıcı oluştur
        new_user = User(
            username=data['username'],
            first_name=data['first_name'],
            last_name=data['last_name'],
            email=data.get('email'),
            role=role_mapping[data['role']],
            representative_code=rep_code,
            is_active=data.get('is_active', True)
        )
        # Departman ataması (opsiyonel)
        if data.get('department_id'):
            new_user.department_id = int(data['department_id'])
        
        # Şifreyi hash'le
        new_user.set_password(data['password'])
        
        db.session.add(new_user)
        db.session.commit()
        
        # Activity log ekle
        activity_log = ActivityLog(
            user_id=current_user.id,
            action='user_created',
            description=f'Yeni kullanıcı oluşturuldu: {new_user.username} ({new_user.role.value})'
        )
        db.session.add(activity_log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kullanıcı başarıyla oluşturuldu',
            'user': {
                'id': new_user.id,
                'username': new_user.username,
                'first_name': new_user.first_name,
                'last_name': new_user.last_name,
                'email': new_user.email,
                'role': new_user.role.value,
                'representative_code': new_user.representative_code,
                'is_active': new_user.is_active
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@api.route('/users/<int:user_id>', methods=['PUT', 'DELETE'])
@admin_required
def manage_user(user_id):
    """Kullanıcı bilgilerini güncelle veya sil"""
    try:
        user = User.query.get_or_404(user_id)
        
        if request.method == 'DELETE':
            # Silme: varsayılan strateji soft-delete (giriş engeli + listelerden gizle)
            # Kendini silmeye çalışıyorsa engelle
            if user.id == current_user.id:
                return jsonify({'success': False,'error': 'Kendinizi silemezsiniz'}), 400
            # Son admin güvenliği
            if user.role == UserRole.ADMIN:
                admin_count = User.query.filter_by(role=UserRole.ADMIN, is_active=True).count()
                if admin_count <= 1:
                    return jsonify({'success': False,'error': 'Son aktif admin silinemez'}), 400
            # Soft delete uygula
            user.is_active = False
            try:
                # benzersiz kullanıcı adı için ek
                user.username = f"{user.username}_deleted_{user.id}"
            except Exception:
                pass
            user.email = None
            user.representative_code = None
            db.session.commit()
            # Logla
            try:
                log = ActivityLog(user_id=current_user.id, action='user_soft_deleted', description=f'Kullanıcı pasif edildi: {user.id}')
                db.session.add(log)
                db.session.commit()
            except Exception:
                db.session.rollback()
            return jsonify({'success': True,'message': 'Kullanıcı pasif edildi (giriş yapamaz)'}), 200
        
        # PUT işlemi
        data = request.get_json()
        
        # Kullanıcı adı değişikliği kontrolü
        if data.get('username') and data['username'] != user.username:
            if User.query.filter_by(username=data['username']).first():
                return jsonify({
                    'success': False,
                    'error': 'Bu kullanıcı adı zaten kullanılıyor'
                }), 400
        
        # Email değişikliği kontrolü
        if data.get('email') and data['email'] != user.email:
            if User.query.filter_by(email=data['email']).first():
                return jsonify({
                    'success': False,
                    'error': 'Bu email adresi zaten kullanılıyor'
                }), 400
        
        # Temsilci kodu kontrolü
        if data.get('representative_code'):
            rep_code = data['representative_code'].strip() if data['representative_code'] else None
            if rep_code and rep_code != user.representative_code:
                if User.query.filter_by(representative_code=rep_code).first():
                    return jsonify({
                        'success': False,
                        'error': 'Bu temsilci kodu zaten kullanılıyor'
                    }), 400
        
        # Bilgileri güncelle
        if 'username' in data:
            user.username = data['username']
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'email' in data:
            user.email = data['email']
        if 'role' in data:
            # Yeni üçlü rol sistemi: ADMIN, DEPARTMENT_MANAGER, USER
            role_mapping = {
                'ADMIN': UserRole.ADMIN,
                'DEPARTMENT_MANAGER': UserRole.DEPARTMENT_MANAGER,
                'USER': UserRole.USER
            }
            if data['role'] not in role_mapping:
                return jsonify({'success': False, 'error': f'Geçersiz rol: {data["role"]}'}), 400
            user.role = role_mapping[data['role']]
            # Temsilci kodu opsiyonel alan, rol bağımsız saklanabilir
        if 'representative_code' in data:
            # Boş string'leri None olarak değiştir (UNIQUE constraint için)
            rep_code = data['representative_code']
            user.representative_code = rep_code if rep_code and rep_code.strip() else None
        if 'is_active' in data:
            user.is_active = data['is_active']
        
        # Şifre değişikliği
        if data.get('password'):
            user.set_password(data['password'])
        
        db.session.commit()
        
        # Activity log ekle
        activity_log = ActivityLog(
            user_id=current_user.id,
            action='user_updated',
            description=f'Kullanıcı güncellendi: {user.username}'
        )
        db.session.add(activity_log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Kullanıcı başarıyla güncellendi',
            'user': {
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'role': user.role.value,
                'representative_code': user.representative_code,
                'is_active': user.is_active
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500





@api.route('/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_user_password(user_id):
    """Kullanıcı şifresini sıfırla"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        if not data.get('new_password'):
            return jsonify({
                'success': False,
                'error': 'Yeni şifre gereklidir'
            }), 400
        
        user.set_password(data['new_password'])
        db.session.commit()
        
        # Activity log ekle
        activity_log = ActivityLog(
            user_id=current_user.id,
            action='password_reset',
            description=f'Kullanıcı şifresi sıfırlandı: {user.username}'
        )
        db.session.add(activity_log)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Şifre başarıyla sıfırlandı'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# Kayıt Devir (Satış/İade) Endpointi
@api.route('/users/<int:from_user_id>/reassign-records', methods=['POST'])
@admin_or_department_manager_required
def reassign_user_records(from_user_id: int):
    """Bir kullanıcının tüm satış ve iade kayıtlarını başka bir kullanıcıya devret.
    - Admin: herkes arasında devir yapabilir
    - DM: kendi departmanı içindeki kullanıcılar arasında devir yapabilir
    Body: { "to_user_id": 123 }
    """
    try:
        data = request.get_json() or {}
        to_user_id = data.get('to_user_id')
        if not to_user_id:
            return jsonify({'success': False, 'error': 'Hedef kullanıcı (to_user_id) gerekli'}), 400

        if from_user_id == to_user_id:
            return jsonify({'success': False, 'error': 'Aynı kullanıcıya devir yapılamaz'}), 400

        from_user = User.query.get_or_404(from_user_id)
        to_user = User.query.get_or_404(to_user_id)

        # Yetki/scope kontrolü: DM sadece kendi departmanı içinde işlem yapabilir
        if current_user.is_department_manager() and (
            from_user.department_id != current_user.department_id or to_user.department_id != current_user.department_id
        ):
            return jsonify({'success': False, 'error': 'Sadece kendi departmanınızdaki kullanıcılar arasında devir yapabilirsiniz'}), 403

        # Devir işlemleri
        Sales.query.filter_by(representative_id=from_user_id).update({Sales.representative_id: to_user_id})
        Returns.query.filter_by(representative_id=from_user_id).update({Returns.representative_id: to_user_id})
        db.session.commit()

        log_activity('records_reassign', f'Kayıt devri: {from_user.username} -> {to_user.username}')
        return jsonify({'success': True, 'message': 'Satış ve iade kayıtları devredildi'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500